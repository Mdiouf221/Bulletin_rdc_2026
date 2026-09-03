"""
generer_annexe_b_visuels.py
============================
Génère les visuels statistiques statiques de l'annexe B (fiches institutionnelles)
directement à partir des données ESS en base — sans navigateur, sans simulation de clics.

Pour chaque institution présente en base, régénère automatiquement :
  1. Les graphiques (cotisants, bénéficiaires, dépenses, dépense moyenne par bénéficiaire,
     recettes, contribution moyenne) exportés en PNG via les MÊMES fonctions Plotly que
     le tableau de bord interactif (build_fig_institution_*) — aucune duplication de la
     logique visuelle : un changement de style dans le dashboard se répercute ici aussi.
  2. Un tableau Markdown natif « Régimes gérés » (description structurée par régime).
  3. Des camemberts « Répartition par sexe » (cotisants et bénéficiaires cumulés), un
     jeu de 2 camemberts par année, exportés en PNG via Plotly/kaleido.
  4. Un tableau Markdown natif « Données détaillées » (par régime et année).

Tous les régimes de l'institution sont inclus. Les corrections enregistrées dans
10_output/questionnaire_data.json (Q1, Q1b, Q2 et Q4) sont appliquées aux graphiques.
Les institutions sans formulaire enregistré conservent leurs données ESS brutes.

Couverture 2019-2025 : le bulletin couvre les années 2019 à 2025. Toute année de
cette plage sans donnée ESS apparaît explicitement comme un repère [N/D] (tableaux)
ou un placeholder visuel (graphiques, camemberts) — jamais silencieusement omise, et
jamais confondue avec une valeur nulle réelle.

Le contenu généré est injecté dans 04_annexes/annexe_B_fiches_institutionnelles.md entre
les marqueurs :
    <!-- AUTO_GENERE:<INSTITUTION>:DEBUT -->
    ...
    <!-- AUTO_GENERE:<INSTITUTION>:FIN -->
Le texte rédigé manuellement en dehors de ces marqueurs (cadre juridique, réformes, etc.)
n'est jamais modifié.

Usage :
    py 09_scripts/generer_annexe_b_visuels.py
"""

import copy
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
except ImportError:
    sys.exit("Plotly requis : py -m pip install plotly")

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:
    sys.exit("openpyxl requis : py -m pip install openpyxl")

try:
    import kaleido
except ImportError:
    kaleido = None

from visualiser_regimes import (  # noqa: E402
    load_all, DB_PATH, NOM_COURT, NOM_INSTITUTION,
    build_fig_institution_cotisants,
    build_fig_institution_beneficiaires,
    build_fig_institution_depenses,
    build_fig_institution_depense_par_beneficiaire,
    build_fig_institution_recettes,
    build_fig_institution_contribution,
    build_institution_detail_table,
)

WORKSPACE = SCRIPT_DIR.parent
ILLUSTRATIONS_DIR = WORKSPACE / "04_annexes" / "illustrations"
ANNEXE_B_MD = WORKSPACE / "04_annexes" / "annexe_B_fiches_institutionnelles.md"
QUESTIONNAIRE_DATA_FILE = WORKSPACE / "10_output" / "questionnaire_data.json"
ANNEXE_B_XLSX_FILE = WORKSPACE / "10_output" / "annexe_B_graphiques_par_institution.xlsx"
ANNEXE_B_WORD_DATA_FILE = WORKSPACE / "10_output" / "annexe_B_graphiques_word.json"

CHART_WIDTH = 1000
CHART_HEIGHT = 520
CHART_SCALE = 2

# Le bulletin couvre 2019-2025 : toutes les tables/graphiques doivent afficher un
# repère pour chacune de ces années, même sans donnée ESS (placeholder [N/D]),
# afin de ne jamais confondre une donnée manquante avec une valeur nulle réelle.
BULLETIN_YEARS = list(range(2019, 2026))

SEX_PIE_COLORS = {
    "Hommes": "#2e78c8",
    "Femmes": "#d4487a",
    "Non identifié": "#a9b4c0",
}
# Disposition compacte : plusieurs années par ligne, un seul visuel par institution
# (au lieu d'une image pleine largeur par année) — réduit fortement l'espace vertical.
SEX_PIE_YEARS_PER_ROW = 3
SEX_PIE_CELL_WIDTH = 150   # largeur par pie (px, avant scale)
SEX_PIE_ROW_HEIGHT = 175   # hauteur par ligne d'années (px, avant scale)
SEX_PIE_TOP_MARGIN = 36
SEX_PIE_BOTTOM_MARGIN = 24
SEX_PIE_SCALE = 2

# (clé fichier, libellé, fonction builder — même source que le dashboard interactif)
CHART_SPECS = [
    ("cotisants", "Cotisants actifs", build_fig_institution_cotisants),
    ("beneficiaires", "Bénéficiaires", build_fig_institution_beneficiaires),
    ("depenses", "Dépenses totales", build_fig_institution_depenses),
    ("depense_par_beneficiaire", "Dépense moyenne par bénéficiaire", build_fig_institution_depense_par_beneficiaire),
    ("recettes", "Recettes totales", build_fig_institution_recettes),
    ("contribution", "Contribution moyenne", build_fig_institution_contribution),
]

CHART_DEFAULTS = {
    "cotisants": {
        "unite": "personnes",
        "type": "courbe",
    },
    "beneficiaires": {
        "unite": "personnes",
        "type": "courbe",
    },
    "depenses": {
        "unite": "Mds CDF",
        "type": "barres",
    },
    "depense_par_beneficiaire": {
        "unite": "k CDF / bénéficiaire",
        "type": "courbe",
    },
    "recettes": {
        "unite": "Mds CDF",
        "type": "barres",
    },
    "contribution": {
        "unite": "k CDF / cotisant",
        "type": "courbe",
    },
    "sexe": {
        "unite": "personnes",
        "type": "colonnes empilées (100%)",
    },
}

# Fonctions qui acceptent un paramètre sex_mode (les autres n'en ont pas besoin)
_SEX_MODE_BUILDERS = {build_fig_institution_cotisants, build_fig_institution_beneficiaires}


def load_questionnaire_data(path: Path = QUESTIONNAIRE_DATA_FILE) -> dict:
    """Charge les paramètres institutionnels utilisés par le tableau de bord."""
    if not path.exists():
        raise FileNotFoundError(
            f"Paramètres institutionnels introuvables : {path}. "
            "Enregistrer d'abord le formulaire du tableau de bord."
        )
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Lecture des paramètres institutionnels impossible : {exc}") from exc
    if not isinstance(data, dict):
        raise ValueError("questionnaire_data.json doit contenir un objet par institution.")
    return data


def _fusion_components(answers: dict) -> list[list[str]]:
    """Construit les composantes connexes des paires de régimes cochées."""
    edges = []
    for key, value in answers.items():
        if value != "true":
            continue
        parts = key.split("__")
        if len(parts) == 2:
            edges.append((parts[0], parts[1]))
    if not edges:
        return []

    parent: dict[str, str] = {}

    def find(item: str) -> str:
        parent.setdefault(item, item)
        if parent[item] != item:
            parent[item] = find(parent[item])
        return parent[item]

    for left, right in edges:
        parent[find(left)] = find(right)

    components: dict[str, list[str]] = {}
    for regime in {item for edge in edges for item in edge}:
        components.setdefault(find(regime), []).append(regime)
    return [sorted(regimes) for regimes in components.values() if len(regimes) > 1]


def _merge_population_traces(traces: list[dict], answers: dict) -> list[dict]:
    """Déduplique les populations partagées en retenant le maximum annuel."""
    components = _fusion_components(answers)
    if not components:
        return copy.deepcopy(traces)

    result = []
    processed: set[str] = set()
    for trace in traces:
        regime = trace.get("legendgroup")
        component = next((group for group in components if regime in group), None)
        if component is None:
            result.append(copy.deepcopy(trace))
            continue

        group_key = "__".join(component)
        if group_key in processed:
            continue
        processed.add(group_key)
        members = [item for item in traces if item.get("legendgroup") in component]
        if not members:
            continue

        merged = copy.deepcopy(members[0])
        years = sorted(
            {year for member in members for year in member.get("x", [])},
            key=lambda value: str(value),
        )
        merged["x"] = years
        merged["y"] = []
        for year in years:
            values = []
            for member in members:
                member_years = member.get("x", [])
                if year not in member_years:
                    continue
                value = member.get("y", [])[member_years.index(year)]
                if value is not None:
                    values.append(value)
            merged["y"].append(max(values) if values else None)
        merged["name"] = " + ".join(NOM_COURT.get(code, code) for code in component)
        merged["legendgroup"] = group_key
        merged.pop("stackgroup", None)
        merged["fill"] = "none"
        result.append(merged)
    return result


def _apply_q4_conversion(traces: list[dict], settings: dict) -> list[dict]:
    """Convertit les bénéficiaires selon les unités et coefficients Q4."""
    units = settings.get("Q4") or {}
    coefficients_by_regime = settings.get("Q4_coefficients") or {}
    converted = copy.deepcopy(traces)
    for trace in converted:
        regime = trace.get("legendgroup")
        unit = units.get(regime)
        if not unit or unit == "enfant":
            continue
        coefficients = coefficients_by_regime.get(regime) or {}
        for index, year in enumerate(trace.get("x", [])):
            values = trace.get("y", [])
            if index >= len(values) or values[index] is None:
                continue
            coefficient_data = coefficients.get(str(year)) or coefficients.get(year) or coefficients.get("default")
            coefficient = coefficient_data.get("value") if isinstance(coefficient_data, dict) else None
            if coefficient not in (None, 0):
                values[index] *= coefficient
    return converted


def _q2_components(q2_data: dict, metric_type: str, year) -> list[list[str]]:
    suffix = f"_{metric_type}_{year}"
    answers = {}
    for key, regimes in q2_data.items():
        if not key.startswith("Q2_") or not key.endswith(suffix) or not isinstance(regimes, list):
            continue
        regime = key[3:-len(suffix)]
        for other in regimes:
            answers[f"{regime}__{other}"] = "true"
    return _fusion_components(answers)


def _merge_finance_traces(
    traces: list[dict],
    q2_data: dict,
    metric_type: str,
) -> tuple[list[dict], list[str]]:
    """Déduplique les montants partagés de Q2, année par année."""
    result = copy.deepcopy(traces)
    merged_by_group: dict[str, dict] = {}
    warnings = []
    years = sorted({str(year) for trace in result for year in trace.get("x", [])})
    for year in years:
        for component in _q2_components(q2_data, metric_type, year):
            points = []
            for trace in result:
                if trace.get("legendgroup") not in component:
                    continue
                index = next(
                    (i for i, value in enumerate(trace.get("x", [])) if str(value) == year),
                    None,
                )
                if index is None:
                    continue
                value = trace.get("y", [])[index]
                if value is not None:
                    points.append((trace, index, value))
            if len(points) <= 1:
                continue

            values = [float(value) for _, _, value in points]
            names = list(dict.fromkeys(str(trace.get("name") or "") for trace, _, _ in points))
            if max(values) - min(values) > max(1, abs(max(values)) * 1e-9):
                warnings.append(f"{year} — {metric_type} : valeurs divergentes pour {' + '.join(names)}")

            group_key = "__".join(sorted(trace.get("legendgroup") for trace, _, _ in points))
            if group_key not in merged_by_group:
                merged = copy.deepcopy(points[0][0])
                merged["name"] = " + ".join(names)
                merged["legendgroup"] = f"Q2_{metric_type}_{group_key}"
                merged["x"] = []
                merged["y"] = []
                merged_by_group[group_key] = merged
            merged_by_group[group_key]["x"].append(points[0][0]["x"][points[0][1]])
            merged_by_group[group_key]["y"].append(max(values))
            for trace, index, _ in points:
                trace["y"][index] = None

    remaining = [
        trace for trace in result
        if any(value is not None for value in trace.get("y", []))
    ]
    return remaining + list(merged_by_group.values()), warnings


def apply_questionnaire_to_figure(fig, chart_key: str, settings: dict):
    """Applique à une figure statique les mêmes règles que le formulaire interactif."""
    if fig is None or not settings:
        return fig, []

    figure = fig.to_dict()
    traces = figure.get("data", [])
    warnings = []
    if chart_key == "cotisants":
        traces = _merge_population_traces(traces, settings.get("Q1") or {})
    elif chart_key == "beneficiaires":
        traces = _apply_q4_conversion(traces, settings)
        traces = _merge_population_traces(traces, settings.get("Q1b") or {})
    elif chart_key in {"depenses", "depense_par_beneficiaire"}:
        traces, warnings = _merge_finance_traces(traces, settings.get("Q2") or {}, "depenses")
    elif chart_key in {"recettes", "contribution"}:
        traces, warnings = _merge_finance_traces(traces, settings.get("Q2") or {}, "recettes")
    return go.Figure(data=traces, layout=figure.get("layout", {})), warnings


def _write_image_with_retry(fig, out_path: Path, *, width: int, height: int, scale: int,
                            retries: int = 2, delay: float = 1.5) -> bool:
    """Exporte une figure Plotly en PNG, avec ré-essais en cas d'échec transitoire du
    sous-processus navigateur de kaleido (peut survenir lors de lots de nombreux exports)."""
    last_exc = None
    for attempt in range(retries + 1):
        try:
            fig.write_image(str(out_path), width=width, height=height, scale=scale)
            return True
        except Exception as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(delay)
    print(f"    [AVERT] Export image '{out_path.name}' échoué après {retries + 1} tentative(s) : {last_exc}")
    return False


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_year(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _has_true(values: dict | None) -> bool:
    if not isinstance(values, dict):
        return False
    return any(str(v).lower() == "true" for v in values.values())


def _has_q2_for_metric(q2_data: dict | None, metric: str) -> bool:
    if not isinstance(q2_data, dict):
        return False
    marker = f"_{metric}_"
    return any(
        key.startswith("Q2_") and marker in key and isinstance(val, list) and len(val) > 0
        for key, val in q2_data.items()
    )


def _q4_sources(settings: dict) -> list[str]:
    sources = []
    seen = set()
    q4_coef = settings.get("Q4_coefficients") if isinstance(settings, dict) else {}
    if not isinstance(q4_coef, dict):
        return sources
    for per_regime in q4_coef.values():
        if not isinstance(per_regime, dict):
            continue
        for item in per_regime.values():
            if not isinstance(item, dict):
                continue
            src = _safe_text(item.get("source"))
            if src and src not in seen:
                seen.add(src)
                sources.append(src)
    return sources


def _questionnaire_rules_for_chart(settings: dict, chart_key: str) -> list[str]:
    rules = []
    if chart_key == "cotisants" and _has_true(settings.get("Q1")):
        rules.append("Q1")
    elif chart_key == "beneficiaires":
        if _has_true(settings.get("Q1b")):
            rules.append("Q1b")
        q4_units = settings.get("Q4") if isinstance(settings, dict) else {}
        if isinstance(q4_units, dict) and any(str(unit).lower() != "enfant" for unit in q4_units.values()):
            rules.append("Q4")
    elif chart_key in {"depenses", "depense_par_beneficiaire"} and _has_q2_for_metric(settings.get("Q2"), "depenses"):
        rules.append("Q2")
    elif chart_key in {"recettes", "contribution"} and _has_q2_for_metric(settings.get("Q2"), "recettes"):
        rules.append("Q2")
    elif chart_key == "sexe" and _has_true(settings.get("Q1b")):
        rules.append("Q1b")
    return rules


def _questionnaire_chart_meta(settings: dict, chart_key: str) -> dict:
    if not isinstance(settings, dict):
        return {}
    candidates = [
        settings.get("graph_metadata"),
        settings.get("graphMeta"),
        settings.get("Q5"),
    ]
    for block in candidates:
        if isinstance(block, dict):
            item = block.get(chart_key)
            if isinstance(item, dict):
                return item
    return {}


def _resolve_chart_metadata(
    *,
    institution: str,
    chart_key: str,
    default_title: str,
    settings: dict,
    years: list[int],
) -> dict:
    q_meta = _questionnaire_chart_meta(settings, chart_key)
    defaults = CHART_DEFAULTS.get(chart_key, {})
    title = _safe_text(q_meta.get("titre")) or _safe_text(q_meta.get("title")) or default_title
    unite = _safe_text(q_meta.get("unite")) or _safe_text(q_meta.get("unit")) or _safe_text(defaults.get("unite"))
    source = _safe_text(q_meta.get("source")) or "Base consolidée des ESS OIT/BIT."
    periode = _safe_text(q_meta.get("periode")) or _safe_text(q_meta.get("period"))
    if not periode:
        if years:
            periode = f"{min(years)}–{max(years)}"
        else:
            periode = "2019–2025"
    chart_type = _safe_text(q_meta.get("type")) or _safe_text(defaults.get("type"))
    notes_parts = []
    base_note = _safe_text(q_meta.get("notes"))
    if base_note:
        notes_parts.append(base_note)
    rules = _questionnaire_rules_for_chart(settings, chart_key)
    if rules:
        notes_parts.append("Règles appliquées : " + ", ".join(rules))
    if chart_key == "beneficiaires":
        q4_sources = _q4_sources(settings)
        if q4_sources:
            notes_parts.append("Sources des coefficients Q4 : " + " | ".join(q4_sources))
    notes = " ".join(notes_parts)
    return {
        "institution": institution,
        "chart_key": chart_key,
        "titre": title,
        "unite": unite,
        "source": source,
        "periode": periode,
        "type": chart_type,
        "notes": notes,
        "regles": rules,
    }


def _figure_to_tabular_data(fig, fallback_years: list[int]) -> tuple[list[str], list[list[Any]], list[int]]:
    traces = list(fig.data) if fig is not None else []
    years_set = set(fallback_years or [])
    for trace in traces:
        for x in getattr(trace, "x", []) or []:
            parsed = _parse_year(x)
            if parsed is not None:
                years_set.add(parsed)
    years = sorted(years_set)
    if not years:
        years = list(BULLETIN_YEARS)

    columns = ["Année"]
    series_values: list[tuple[str, dict[int, Any]]] = []
    for idx, trace in enumerate(traces, start=1):
        name = _safe_text(getattr(trace, "name", None)) or f"Série {idx}"
        columns.append(name)
        mapping: dict[int, Any] = {}
        x_vals = list(getattr(trace, "x", []) or [])
        y_vals = list(getattr(trace, "y", []) or [])
        for x, y in zip(x_vals, y_vals):
            year = _parse_year(x)
            if year is None:
                continue
            mapping[year] = y
        series_values.append((name, mapping))

    rows: list[list[Any]] = []
    for year in years:
        row = [year]
        for _name, mapping in series_values:
            row.append(mapping.get(year))
        rows.append(row)
    return columns, rows, years


def _build_chart_payload(
    *,
    institution: str,
    chart_key: str,
    chart_label: str,
    fig,
    settings: dict,
    fallback_years: list[int],
    image_filename: str | None,
) -> dict:
    if fig is not None:
        columns, rows, years = _figure_to_tabular_data(fig, fallback_years)
        layout_title = getattr(getattr(fig, "layout", None), "title", None)
        title = _safe_text(getattr(layout_title, "text", None))
    else:
        years = sorted(set(fallback_years) or set(BULLETIN_YEARS))
        columns = ["Année"]
        rows = [[year] for year in years]
        title = f"{NOM_INSTITUTION.get(institution, institution)} — {chart_label}"

    metadata = _resolve_chart_metadata(
        institution=institution,
        chart_key=chart_key,
        default_title=title,
        settings=settings,
        years=years,
    )
    metadata.update(
        {
            "label": chart_label,
            "image_filename": image_filename,
            "columns": columns,
            "rows": rows,
        }
    )
    return metadata


def generate_charts(
    regimes: list[dict],
    institution: str,
    settings: dict,
) -> tuple[list[tuple[str, str, str | None]], list[dict]]:
    """Génère les graphiques PNG d'une institution.

    Retourne toujours une entrée par graphique défini dans CHART_SPECS, dans l'ordre
    (key, label, filename_ou_None) — même quand le graphique ne peut pas être généré
    (donnée insuffisante), afin que l'emplacement reste visible dans la grille (avec
    un repère [N/D]) plutôt que de disparaître silencieusement.
    """
    ILLUSTRATIONS_DIR.mkdir(parents=True, exist_ok=True)
    inst_years = {r["annee"] for r in regimes if r["institution"] == institution and r.get("annee") is not None}
    x_min = min([*BULLETIN_YEARS, *inst_years]) if inst_years else BULLETIN_YEARS[0]
    x_max = max([*BULLETIN_YEARS, *inst_years]) if inst_years else BULLETIN_YEARS[-1]

    results: list[tuple[str, str, str | None]] = []
    payloads: list[dict] = []
    fallback_years = sorted(set(BULLETIN_YEARS) | {int(year) for year in inst_years if year is not None})
    for key, label, builder in CHART_SPECS:
        fig = None
        try:
            if builder in _SEX_MODE_BUILDERS:
                fig = builder(regimes, institution, "all")
            else:
                fig = builder(regimes, institution)
            fig, warnings = apply_questionnaire_to_figure(fig, key, settings)
            for warning in warnings:
                print(f"    [AVERT] Paramètres Q2 ({institution}) : {warning}")
        except Exception as exc:
            print(f"    [AVERT] Graphique '{label}' ({institution}) : {exc}")
            results.append((key, label, None))
            payloads.append(
                _build_chart_payload(
                    institution=institution,
                    chart_key=key,
                    chart_label=label,
                    fig=None,
                    settings=settings,
                    fallback_years=fallback_years,
                    image_filename=None,
                )
            )
            continue

        if fig is None or not getattr(fig, "data", None):
            results.append((key, label, None))
            payloads.append(
                _build_chart_payload(
                    institution=institution,
                    chart_key=key,
                    chart_label=label,
                    fig=None,
                    settings=settings,
                    fallback_years=fallback_years,
                    image_filename=None,
                )
            )
            continue

        if institution == "TRESOR" and key == "cotisants":
            fig.update_layout(title_text="Personnes potentiellement couvertes (estimation)")

        # Fixe la plage de l'axe des années sur toute la période du bulletin (2019-2025,
        # étendue si des données réelles existent au-delà) : les années sans donnée ESS
        # restent visibles comme un espace vide sur la série, au lieu d'être masquées
        # par le zoom automatique de Plotly sur les seules années présentes.
        fig.update_xaxes(range=[x_min - 0.5, x_max + 0.5])

        filename = f"annexe_B_{institution}_{key}.png"
        out_path = ILLUSTRATIONS_DIR / filename
        try:
            ok = _write_image_with_retry(
                fig,
                out_path,
                width=CHART_WIDTH,
                height=CHART_HEIGHT,
                scale=CHART_SCALE,
            )
        except Exception as exc:
            print(f"    [AVERT] Export image '{label}' ({institution}) échoué : {exc}")
            results.append((key, label, None))
            payloads.append(
                _build_chart_payload(
                    institution=institution,
                    chart_key=key,
                    chart_label=label,
                    fig=fig,
                    settings=settings,
                    fallback_years=fallback_years,
                    image_filename=None,
                )
            )
            continue
        image_filename = filename if ok else None
        results.append((key, label, image_filename))
        payloads.append(
            _build_chart_payload(
                institution=institution,
                chart_key=key,
                chart_label=label,
                fig=fig,
                settings=settings,
                fallback_years=fallback_years,
                image_filename=image_filename,
            )
        )
    return results, payloads


def build_charts_markdown(
    chart_entries: list[tuple[str, str, str | None]],
    institution: str,
) -> str:
    """Dispose les graphiques en grille 2 colonnes (même logique que le dashboard).

    Un emplacement sans graphique disponible affiche un repère [N/D] plutôt que
    d'être omis, afin que l'absence de donnée reste visible dans la mise en page.
    """
    if not chart_entries:
        return "*Aucun graphique disponible (données insuffisantes pour cette institution).*\n"

    def _cell(entry: tuple[str, str, str | None] | None) -> str:
        if entry is None:
            return '<td style="width:50%; padding:4px;"></td>'
        _key, label, filename = entry
        if filename:
            return (
                '<td style="width:50%; padding:4px;">'
                f'<img src="/files/04_annexes/illustrations/{filename}" style="width:100%; height:auto;">'
                "</td>"
            )
        return (
            '<td style="width:50%; padding:4px;">'
            '<div style="min-height:180px; display:flex; align-items:center; justify-content:center; '
            'border:1px dashed #cbd5e0; border-radius:6px; color:#888; font-size:0.9em; text-align:center;">'
            f"[N/D] — {label}</div></td>"
        )

    rows_html = []
    for i in range(0, len(chart_entries), 2):
        pair = chart_entries[i:i + 2]
        cells = _cell(pair[0]) + (_cell(pair[1]) if len(pair) > 1 else _cell(None))
        rows_html.append(f"<tr>{cells}</tr>")

    return (
        '<table style="width:100%; border-collapse:collapse;">\n'
        + "\n".join(rows_html)
        + "\n</table>\n"
    )


def build_regime_table_markdown(regime_meta_inst: dict) -> str:
    """Tableau Markdown natif décrivant chaque régime de l'institution (dernière version connue)."""
    regime_codes = sorted(regime_meta_inst.keys())
    if not regime_codes:
        return "*Aucun régime documenté pour cette institution.*\n"

    lines = [
        "| Régime | Type de financement | Caractère | Gestion | Fonctions couvertes |",
        "|---|---|---|---|---|",
    ]
    for rc in regime_codes:
        meta = regime_meta_inst[rc]
        versions = meta.get("versions") or []
        if not versions:
            continue
        latest = versions[-1]
        nom = latest.get("nom_regime") or NOM_COURT.get(rc, rc)
        fonctions = latest.get("fonctions_oit") or []
        fonctions_txt = "; ".join(fonctions) if fonctions else "—"
        lines.append(
            f"| {nom} | {latest.get('type_financement') or '—'} | "
            f"{latest.get('caractere') or '—'} | {latest.get('gestion') or '—'} | "
            f"{fonctions_txt} |"
        )
    return "\n".join(lines) + "\n"


def _sex_pie_slice(h: float, f: float, ni: float, total: float):
    """Retourne (labels, valeurs, couleurs) pour un camembert H/F/non identifié.

    Si aucune donnée n'est disponible pour l'année (total = 0), affiche un
    unique secteur neutre plutôt qu'un camembert vide. Les catégories à valeur
    nulle sont exclues pour éviter des étiquettes « 0 % » parasites.
    """
    if not total:
        return ["Non disponible"], [1], ["#d9dee5"]
    labels_all = ["Hommes", "Femmes", "Non identifié"]
    values_all = [h, f, ni]
    filtered = [
        (label, value, SEX_PIE_COLORS[label])
        for label, value in zip(labels_all, values_all)
        if value and value > 0
    ]
    if not filtered:
        return ["Non disponible"], [1], ["#d9dee5"]
    labels, values, colors = zip(*filtered)
    return list(labels), list(values), list(colors)


def build_sex_pie_grid_figure(
    year_data: dict[int, dict],
    *,
    large_format: bool = False,
    contributors_label: str = "Cotisants",
):
    """Construit une grille compacte de camemberts H/F/non identifié — plusieurs années
    par ligne (SEX_PIE_YEARS_PER_ROW), un seul visuel pour toutes les années de
    l'institution (au lieu d'une image pleine largeur par année).

    Chaque année occupe 2 colonnes (cotisants, bénéficiaires), identifiées directement
    sous leur camembert. Seule l'année est rappelée au-dessus de chaque paire.
    """
    years = sorted(year_data.keys())
    n_years = len(years)
    if n_years == 0:
        return None

    years_per_row = 2 if large_format else SEX_PIE_YEARS_PER_ROW
    cell_width = 220 if large_format else SEX_PIE_CELL_WIDTH
    row_height = 240 if large_format else SEX_PIE_ROW_HEIGHT
    cols_per_row = min(n_years, years_per_row)
    n_cols = cols_per_row * 2
    n_rows = -(-n_years // cols_per_row)  # ceil division

    specs = []
    for r in range(n_rows):
        row_specs = []
        for c in range(cols_per_row):
            idx = r * cols_per_row + c
            if idx < n_years:
                row_specs.extend([{"type": "domain"}, {"type": "domain"}])
            else:
                row_specs.extend([None, None])
        specs.append(row_specs)

    fig = make_subplots(
        rows=n_rows, cols=n_cols, specs=specs,
        horizontal_spacing=0.03,
        vertical_spacing=0.22 if n_rows > 1 else 0.05,
    )

    label_font = dict(size=24 if large_format else 11, family='-apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif', color='#2c5282')
    # Décalage de l'étiquette "année" en pixels fixes (converti en fraction du domaine).
    # Un décalage exprimé directement en fraction (ex. +0.05) grandirait en pixels avec
    # le nombre de lignes (le domaine s'agrandit), au risque de faire déborder
    # l'étiquette de la marge haute — d'où ce calcul dépendant de n_rows.
    domain_height_px = n_rows * row_height
    year_label_offset = 18.0 / domain_height_px
    category_label_offset = 8.0 / domain_height_px
    category_font = dict(
        size=22 if large_format else 10,
        family='-apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif',
        color='#4a5568',
    )

    for r in range(n_rows):
        for c in range(cols_per_row):
            idx = r * cols_per_row + c
            if idx >= n_years:
                continue
            year = years[idx]
            row = year_data[year]
            cot_total, cot_h, cot_f = row["cotisants_total"], row["cotisants_h"], row["cotisants_f"]
            cot_ni = max(cot_total - cot_h - cot_f, 0.0)
            ben_total, ben_h, ben_f = row["beneficiaires_total"], row["beneficiaires_h"], row["beneficiaires_f"]
            ben_ni = max(ben_total - ben_h - ben_f, 0.0)

            cot_labels, cot_values, cot_colors = _sex_pie_slice(cot_h, cot_f, cot_ni, cot_total)
            ben_labels, ben_values, ben_colors = _sex_pie_slice(ben_h, ben_f, ben_ni, ben_total)

            row_idx = r + 1
            col_cot = c * 2 + 1
            col_ben = c * 2 + 2

            fig.add_trace(go.Pie(
                labels=cot_labels, values=cot_values, hole=0.5,
                marker=dict(colors=cot_colors, line=dict(color="#ffffff", width=1)),
                textinfo="percent", textfont=dict(size=22 if large_format else 9), showlegend=False,
                hovertemplate="%{label} : %{value:,.0f} (%{percent})<extra></extra>",
            ), row=row_idx, col=col_cot)

            fig.add_trace(go.Pie(
                labels=ben_labels, values=ben_values, hole=0.5,
                marker=dict(colors=ben_colors, line=dict(color="#ffffff", width=1)),
                textinfo="percent", textfont=dict(size=22 if large_format else 9), showlegend=False,
                hovertemplate="%{label} : %{value:,.0f} (%{percent})<extra></extra>",
            ), row=row_idx, col=col_ben)

            dom_cot = fig.get_subplot(row_idx, col_cot)
            dom_ben = fig.get_subplot(row_idx, col_ben)
            x_center = (dom_cot.x[0] + dom_ben.x[1]) / 2
            fig.add_annotation(
                x=x_center, y=dom_cot.y[1] + year_label_offset, xref="paper", yref="paper",
                text=f"<b>{year}</b>", showarrow=False, font=label_font,
                xanchor="center", yanchor="bottom",
            )
            for domain, text in ((dom_cot, contributors_label), (dom_ben, "Bénéficiaires")):
                fig.add_annotation(
                    x=(domain.x[0] + domain.x[1]) / 2,
                    y=domain.y[0] - category_label_offset,
                    xref="paper",
                    yref="paper",
                    text=text,
                    showarrow=False,
                    font=category_font,
                    xanchor="center",
                    yanchor="top",
                )

    fig.update_layout(
        height=SEX_PIE_TOP_MARGIN + SEX_PIE_BOTTOM_MARGIN + n_rows * row_height,
        width=max(500, n_cols * cell_width + 40),
        showlegend=False,
        plot_bgcolor="#ffffff",
        paper_bgcolor="#ffffff",
        margin=dict(t=SEX_PIE_TOP_MARGIN, b=SEX_PIE_BOTTOM_MARGIN, l=10, r=10),
        font=dict(family='-apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif', size=22 if large_format else 12, color='#4a5568'),
    )
    return fig


def _aggregate_sex_data(regimes: list[dict], institution: str, settings: dict) -> dict[int, dict]:
    """Agrège les données sexuées comme l'aperçu interactif, Q1b compris."""
    fields = (
        "cotisants_total", "cotisants_h", "cotisants_f",
        "beneficiaires_total", "beneficiaires_h", "beneficiaires_f",
    )
    by_year_regime: dict[int, dict[str, dict]] = {}
    for row in regimes:
        if row["institution"] != institution or row.get("annee") is None:
            continue
        year = row["annee"]
        regime = row["regime_code"]
        acc = by_year_regime.setdefault(year, {}).setdefault(
            regime, {field: 0.0 for field in fields},
        )
        for field in fields:
            value = row.get(field)
            if value is not None:
                acc[field] += float(value)

    regime_to_group = {}
    for component in _fusion_components(settings.get("Q1b") or {}):
        group_key = "__".join(component)
        for regime in component:
            regime_to_group[regime] = group_key

    by_year: dict[int, dict] = {
        year: {field: 0.0 for field in fields}
        for year in BULLETIN_YEARS
    }
    for year, regime_rows in by_year_regime.items():
        acc = by_year.setdefault(year, {field: 0.0 for field in fields})
        for row in regime_rows.values():
            for field in ("cotisants_total", "cotisants_h", "cotisants_f"):
                acc[field] += row[field]

        grouped: dict[str, list[dict]] = {}
        for regime, row in regime_rows.items():
            grouped.setdefault(regime_to_group.get(regime, regime), []).append(row)
        for rows in grouped.values():
            representative = max(
                rows,
                key=lambda row: max(
                    row["beneficiaires_total"],
                    row["beneficiaires_h"] + row["beneficiaires_f"],
                ),
            )
            for field in ("beneficiaires_total", "beneficiaires_h", "beneficiaires_f"):
                acc[field] += representative[field]
    return by_year


def _build_sex_payload(
    *,
    institution: str,
    settings: dict,
    by_year: dict[int, dict],
    image_filename: str | None,
) -> dict:
    years = sorted(by_year.keys()) if by_year else list(BULLETIN_YEARS)
    columns = [
        "Année",
        "Cotisants - Hommes",
        "Cotisants - Femmes",
        "Cotisants - Non identifié",
        "Bénéficiaires - Hommes",
        "Bénéficiaires - Femmes",
        "Bénéficiaires - Non identifié",
    ]
    rows = []
    for year in years:
        row = by_year.get(year, {})
        cot_total = float(row.get("cotisants_total") or 0.0)
        cot_h = float(row.get("cotisants_h") or 0.0)
        cot_f = float(row.get("cotisants_f") or 0.0)
        cot_ni = max(cot_total - cot_h - cot_f, 0.0)
        ben_total = float(row.get("beneficiaires_total") or 0.0)
        ben_h = float(row.get("beneficiaires_h") or 0.0)
        ben_f = float(row.get("beneficiaires_f") or 0.0)
        ben_ni = max(ben_total - ben_h - ben_f, 0.0)
        rows.append([year, cot_h, cot_f, cot_ni, ben_h, ben_f, ben_ni])

    default_title = (
        f"{NOM_INSTITUTION.get(institution, institution)} — Répartition par sexe"
        f" ({'personnes potentiellement couvertes estimées' if institution == 'TRESOR' else 'cotisants'} et bénéficiaires)"
    )
    metadata = _resolve_chart_metadata(
        institution=institution,
        chart_key="sexe",
        default_title=default_title,
        settings=settings,
        years=years,
    )
    metadata.update(
        {
            "label": "Répartition par sexe",
            "image_filename": image_filename,
            "columns": columns,
            "rows": rows,
        }
    )
    return metadata


def generate_sex_pie_charts(
    regimes: list[dict],
    institution: str,
    settings: dict,
) -> tuple[list[str], dict]:
    """Génère un unique visuel en grille (cotisants/bénéficiaires par année, plusieurs
    années par ligne) pour l'institution. Retourne les fichiers produits (0 ou 1)."""
    empty_payload = _build_sex_payload(
        institution=institution,
        settings=settings,
        by_year={year: {} for year in BULLETIN_YEARS},
        image_filename=None,
    )
    if not any(r["institution"] == institution for r in regimes):
        return [], empty_payload

    by_year = _aggregate_sex_data(regimes, institution, settings)

    ILLUSTRATIONS_DIR.mkdir(parents=True, exist_ok=True)

    # Nettoyer les anciens fichiers par année (ancien format annexe_B_<INST>_sexe_<annee>.png)
    for old_file in ILLUSTRATIONS_DIR.glob(f"annexe_B_{institution}_sexe_*.png"):
        try:
            old_file.unlink()
        except OSError:
            pass

    try:
        fig = build_sex_pie_grid_figure(
            by_year,
            large_format=False,
            contributors_label=(
                "Personnes couvertes estimées" if institution == "TRESOR" else "Cotisants"
            ),
        )
    except Exception as exc:
        print(f"    [AVERT] Grille camemberts sexe ({institution}) : {exc}")
        return [], _build_sex_payload(
            institution=institution,
            settings=settings,
            by_year=by_year,
            image_filename=None,
        )

    if fig is None:
        return [], _build_sex_payload(
            institution=institution,
            settings=settings,
            by_year=by_year,
            image_filename=None,
        )

    filename = f"annexe_B_{institution}_sexe.png"
    out_path = ILLUSTRATIONS_DIR / filename
    width = fig.layout.width
    height = fig.layout.height
    if not _write_image_with_retry(fig, out_path, width=width, height=height, scale=SEX_PIE_SCALE):
        return [], _build_sex_payload(
            institution=institution,
            settings=settings,
            by_year=by_year,
            image_filename=None,
        )
    return [filename], _build_sex_payload(
        institution=institution,
        settings=settings,
        by_year=by_year,
        image_filename=filename,
    )


def build_sex_charts_markdown(chart_files: list[str], institution: str) -> str:
    """Légende partagée (une seule fois) + image en grille compacte."""
    if not chart_files:
        return "*Aucune donnée de répartition par sexe disponible.*\n"

    legend_items = " &nbsp;&nbsp; ".join(
        '<span style="display:inline-block;width:9px;height:9px;border-radius:50%;'
        f'background:{color};margin-right:4px;"></span>{label}'
        for label, color in SEX_PIE_COLORS.items()
    )
    legend_html = (
        '<p align="center" style="font-size:0.85em; '
        'color:#4a5568; margin-bottom:4px;">'
        f"{legend_items}</p>\n"
    )
    image_html = "\n".join(
        f'<p align="center"><img src="/files/04_annexes/illustrations/{name}" '
        'style="width:100%; height:auto; max-width:620px;"></p>'
        for name in chart_files
    )
    return legend_html + image_html + "\n"


def _sanitize_excel_sheet_name(name: str) -> str:
    clean = re.sub(r'[\[\]\*\?/:\\]', "_", name or "")
    clean = re.sub(r"\s+", "_", clean).strip("_")
    return clean[:31] or "Feuille"


def _unique_sheet_name(base: str, used: set[str]) -> str:
    candidate = _sanitize_excel_sheet_name(base)
    if candidate not in used:
        used.add(candidate)
        return candidate
    idx = 2
    while True:
        suffix = f"_{idx}"
        truncated = _sanitize_excel_sheet_name(candidate[: 31 - len(suffix)] + suffix)
        if truncated not in used:
            used.add(truncated)
            return truncated
        idx += 1


def _append_chart_sheet(ws, payload: dict):
    ws["A1"] = "Institution"
    ws["B1"] = payload.get("institution", "")
    ws["A2"] = "Graphique"
    ws["B2"] = payload.get("chart_key", "")
    ws["A3"] = "Titre"
    ws["B3"] = payload.get("titre", "")
    ws["A4"] = "Unité"
    ws["B4"] = payload.get("unite", "")
    ws["A5"] = "Source"
    ws["B5"] = payload.get("source", "")
    ws["A6"] = "Période"
    ws["B6"] = payload.get("periode", "")
    ws["A7"] = "Type"
    ws["B7"] = payload.get("type", "")
    ws["A8"] = "Notes"
    ws["B8"] = payload.get("notes", "")
    ws["A9"] = "Règles questionnaire"
    ws["B9"] = ", ".join(payload.get("regles", []))

    columns = payload.get("columns") or ["Année"]
    rows = payload.get("rows") or []
    data_start = 11

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=data_start, column=col_idx, value=col_name)
        bold_font = copy.copy(cell.font)
        bold_font.bold = True
        cell.font = bold_font

    for row_idx, row_values in enumerate(rows, start=data_start + 1):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    if len(columns) >= 2:
        for col_idx in range(2, len(columns) + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 18

    if len(columns) <= 1 or not rows:
        return

    data_end_row = data_start + len(rows)
    categories = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end_row)
    chart_key = payload.get("chart_key")

    if chart_key == "sexe" and len(columns) >= 7:
        cot_chart = BarChart()
        cot_chart.type = "col"
        cot_chart.grouping = "percentStacked"
        cot_chart.overlap = 100
        cot_chart.title = "Répartition par sexe — cotisants"
        cot_chart.y_axis.title = "%"
        cot_data = Reference(ws, min_col=2, max_col=4, min_row=data_start, max_row=data_end_row)
        cot_chart.add_data(cot_data, titles_from_data=True)
        cot_chart.set_categories(categories)
        cot_chart.height = 7
        cot_chart.width = 11
        ws.add_chart(cot_chart, "A14")

        ben_chart = BarChart()
        ben_chart.type = "col"
        ben_chart.grouping = "percentStacked"
        ben_chart.overlap = 100
        ben_chart.title = "Répartition par sexe — bénéficiaires"
        ben_chart.y_axis.title = "%"
        ben_data = Reference(ws, min_col=5, max_col=7, min_row=data_start, max_row=data_end_row)
        ben_chart.add_data(ben_data, titles_from_data=True)
        ben_chart.set_categories(categories)
        ben_chart.height = 7
        ben_chart.width = 11
        ws.add_chart(ben_chart, "M14")
        return

    if chart_key in {"depenses", "recettes"}:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
    else:
        chart = LineChart()
        chart.grouping = "standard"

    chart.title = payload.get("titre", payload.get("label", "Graphique"))
    chart.y_axis.title = payload.get("unite", "")
    chart.x_axis.title = "Année"
    chart.height = 8
    chart.width = 24
    data_ref = Reference(ws, min_col=2, max_col=len(columns), min_row=data_start, max_row=data_end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    chart.style = 2
    ws.add_chart(chart, "A14")


def export_annexe_b_workbook(chart_payloads: list[dict]) -> tuple[dict, str]:
    ANNEXE_B_XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    index_ws = wb.create_sheet("index")
    index_headers = [
        "institution",
        "chart_key",
        "sheet_name",
        "png",
        "titre",
        "unite",
        "source",
        "periode",
        "type",
        "notes",
        "regles",
    ]
    for col_idx, header in enumerate(index_headers, start=1):
        cell = index_ws.cell(row=1, column=col_idx, value=header)
        bold_font = copy.copy(cell.font)
        bold_font.bold = True
        cell.font = bold_font

    used_names = {"index"}
    word_payload: dict[str, dict] = {}
    for row_idx, payload in enumerate(chart_payloads, start=2):
        base_sheet = f"{payload.get('institution', 'INST')}_{payload.get('chart_key', 'graph')}"
        sheet_name = _unique_sheet_name(base_sheet, used_names)
        payload["sheet_name"] = sheet_name
        ws = wb.create_sheet(sheet_name)
        _append_chart_sheet(ws, payload)

        index_ws.cell(row=row_idx, column=1, value=payload.get("institution"))
        index_ws.cell(row=row_idx, column=2, value=payload.get("chart_key"))
        index_ws.cell(row=row_idx, column=3, value=sheet_name)
        index_ws.cell(row=row_idx, column=4, value=payload.get("image_filename"))
        index_ws.cell(row=row_idx, column=5, value=payload.get("titre"))
        index_ws.cell(row=row_idx, column=6, value=payload.get("unite"))
        index_ws.cell(row=row_idx, column=7, value=payload.get("source"))
        index_ws.cell(row=row_idx, column=8, value=payload.get("periode"))
        index_ws.cell(row=row_idx, column=9, value=payload.get("type"))
        index_ws.cell(row=row_idx, column=10, value=payload.get("notes"))
        index_ws.cell(row=row_idx, column=11, value=", ".join(payload.get("regles", [])))

        image_filename = payload.get("image_filename")
        if image_filename:
            word_payload[image_filename] = {
                "institution": payload.get("institution"),
                "chart_key": payload.get("chart_key"),
                "sheet_name": sheet_name,
                "titre": payload.get("titre"),
                "unite": payload.get("unite"),
                "source": payload.get("source"),
                "periode": payload.get("periode"),
                "type": payload.get("type"),
                "notes": payload.get("notes"),
                "regles": payload.get("regles", []),
                "columns": payload.get("columns", []),
                "rows": payload.get("rows", []),
            }

    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"):
        index_ws.column_dimensions[col].width = 22

    wb.save(ANNEXE_B_XLSX_FILE)
    workbook_rel = str(ANNEXE_B_XLSX_FILE.relative_to(WORKSPACE)).replace("\\", "/")
    return word_payload, workbook_rel


def export_annexe_b_word_data(chart_payloads: list[dict]):
    charts, workbook_rel = export_annexe_b_workbook(chart_payloads)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "xlsx": workbook_rel,
        "charts": charts,
    }
    ANNEXE_B_WORD_DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    ANNEXE_B_WORD_DATA_FILE.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  XLSX annexe B généré : {ANNEXE_B_XLSX_FILE.relative_to(WORKSPACE)}")
    print(f"  Métadonnées Word annexe B : {ANNEXE_B_WORD_DATA_FILE.relative_to(WORKSPACE)}")


def build_detailed_data_markdown(regimes: list[dict], regime_meta: dict, institution: str) -> str:
    """Tableau Markdown natif — données détaillées par régime et année (mode « tous »).

    Réutilise build_institution_detail_table pour les données réelles (même source
    que le tableau « Données détaillées » du dashboard interactif), puis complète
    avec des lignes [N/D] pour chaque régime connu × chaque année du bulletin
    (2019-2025) sans donnée ESS, afin qu'aucune année ne soit silencieusement omise
    et qu'une absence de donnée ne soit jamais confondue avec une valeur nulle réelle.
    Les noms de régime sont alignés sur ceux du tableau « Régimes gérés » (regime_meta).

    Pour rester compact, les années consécutives sans aucune donnée sont fusionnées en
    une seule ligne « AAAA–AAAA » (cf. boucle de compaction ci-dessous) plutôt que
    répétées une par une : aucune information n'est perdue (l'absence de donnée reste
    explicite), seul le nombre de lignes entièrement vides est réduit.
    """
    table = build_institution_detail_table(regimes, institution, "all")
    inst_regime_meta = regime_meta.get(institution, {})

    if table is None and not inst_regime_meta:
        return "*Aucune donnée détaillée disponible.*\n"

    headers = table["headers"] if table else [
        "Régime", "Année", "Cotisants totaux", "Bénéficiaires totaux",
        "Dépenses totales (Mds CDF)", "Recettes totales (Mds CDF)",
        "Dép. moy./bénéf. (k CDF)", "Rec. moy./cotisant (k CDF)",
    ]
    if institution == "TRESOR":
        headers = [
            "Personnes potentiellement couvertes (estimation)"
            if header == "Cotisants totaux"
            else header.replace("cotisant", "personne couverte estimée")
            for header in headers
        ]
    n_value_cols = len(headers) - 2  # hors « Régime » et « Année »

    # Ordre des régimes : ordre de première apparition dans les données réelles,
    # puis régimes connus (regime_meta) sans aucune donnée ESS, à la fin.
    regime_order: list[str] = []
    real_rows_by_key: dict[tuple[str, int], list[str]] = {}
    real_years_by_regime: dict[str, set[int]] = {}
    if table:
        for item in table["rows"]:
            rc = item["regime_code"]
            year = item["annee"]
            if rc not in regime_order:
                regime_order.append(rc)
            real_rows_by_key[(rc, year)] = item["values"]
            real_years_by_regime.setdefault(rc, set()).add(year)
    for rc in sorted(inst_regime_meta.keys()):
        if rc not in regime_order:
            regime_order.append(rc)

    def regime_label(rc: str) -> str:
        meta = inst_regime_meta.get(rc) or {}
        versions = meta.get("versions") or []
        if versions:
            nom = versions[-1].get("nom_regime")
            if nom:
                return nom
        return NOM_COURT.get(rc, rc)

    lines = [
        "| " + " | ".join(headers) + " |",
        "|" + "|".join("---" for _ in headers) + "|",
    ]
    for rc in regime_order:
        years_for_regime = sorted(set(BULLETIN_YEARS) | real_years_by_regime.get(rc, set()))
        label = regime_label(rc)
        # Compacité : les années consécutives sans aucune donnée ESS sont fusionnées en
        # une seule ligne « AAAA–AAAA | [N/D] | … » plutôt que répétées une par une
        # (ex. trois lignes 100 % [N/D] pour 2023, 2024, 2025 deviennent une seule ligne
        # « 2023–2025 »). Les années avec données réelles restent sur des lignes
        # distinctes, car leurs valeurs diffèrent d'une année à l'autre.
        i = 0
        n_years = len(years_for_regime)
        while i < n_years:
            year = years_for_regime[i]
            values = real_rows_by_key.get((rc, year))
            if values is not None:
                lines.append("| " + " | ".join([label] + values[1:]) + " |")
                i += 1
                continue
            j = i
            while (
                j + 1 < n_years
                and years_for_regime[j + 1] == years_for_regime[j] + 1
                and (rc, years_for_regime[j + 1]) not in real_rows_by_key
            ):
                j += 1
            year_label = str(year) if j == i else f"{year}–{years_for_regime[j]}"
            row = [label, year_label] + ["[N/D]"] * n_value_cols
            lines.append("| " + " | ".join(row) + " |")
            i = j + 1

    return "\n".join(lines) + "\n"


def _institution_caption_label(institution: str) -> str:
    """Sigle court utilisé dans les légendes (ex. « CNSS », « CNSSAP »), distinct du
    libellé long NOM_INSTITUTION utilisé dans les titres des graphiques Plotly."""
    if institution == "TRESOR":
        return 'dispositif budgétaire hors CNSSAP (proxy technique « TRESOR »)'
    return institution


def _locate_section_no_and_counters(md_text: str, institution: str) -> tuple[int, int, int]:
    """Repère le numéro de la fiche institutionnelle (« # B.N ») qui précède le
    marqueur AUTO_GENERE de l'institution, puis compte les légendes « Tableau B.N.x »
    et « Figure B.N.x » déjà présentes dans cette section (tableaux rédigés à la main
    en amont du bloc auto-généré) afin de poursuivre la numérotation sans collision,
    quelle que soit la structure propre à chaque institution (certaines fiches n'ont
    aucun tableau manuel avant le bloc auto-généré, par ex. le Trésor)."""
    marker = f"<!-- AUTO_GENERE:{institution}:DEBUT -->"
    idx = md_text.find(marker)
    if idx == -1:
        return 0, 1, 1
    head_matches = list(re.finditer(r"(?m)^# B\.(\d+)\b", md_text[:idx]))
    if not head_matches:
        return 0, 1, 1
    last = head_matches[-1]
    section_no = int(last.group(1))
    section_text = md_text[last.start():idx]
    n_tables = len(re.findall(rf"Tableau B\.{section_no}\.\d+", section_text))
    n_figures = len(re.findall(rf"Figure B\.{section_no}\.\d+", section_text))
    return section_no, n_tables + 1, n_figures + 1


def build_section_content(
    regimes: list[dict],
    regime_meta: dict,
    institution: str,
    md_text: str,
    settings: dict,
    chart_payloads: list[dict] | None = None,
) -> str:
    label = _institution_caption_label(institution)
    section_no, table_no, figure_no = _locate_section_no_and_counters(md_text, institution)
    family_beneficiaries_footnote = (
        '<span class="footnote">Pour les allocations familiales de la CNSS, le nombre '
        "d'enfants bénéficiaires est estimé en multipliant par 3,17 le nombre de titulaires "
        "de prestations familiales communiqué par la CNSS. Ce facteur correspond au nombre "
        "moyen d'enfants de moins de 20 ans par foyer en RDC en 2013, d'après UN HH Size and "
        "Composition 2019. Il s'agit donc d'une estimation et non d'un décompte administratif "
        "direct d'enfants.</span>"
        if institution == "CNSS" else ""
    )

    chart_entries, chart_payload_entries = generate_charts(regimes, institution, settings)
    if chart_payloads is not None:
        chart_payloads.extend(chart_payload_entries)
    charts_md = build_charts_markdown(chart_entries, institution)
    regimes_md = build_regime_table_markdown(regime_meta.get(institution, {}))
    sex_chart_files, sex_payload = generate_sex_pie_charts(regimes, institution, settings)
    if chart_payloads is not None:
        chart_payloads.append(sex_payload)
    sex_md = build_sex_charts_markdown(sex_chart_files, institution)
    detail_md = build_detailed_data_markdown(regimes, regime_meta, institution)
    tresor_detail_footnote = (
        '\n<span class="footnote">Les valeurs de la colonne « Personnes potentiellement '
        'couvertes (estimation) » ne sont pas des cotisants. Elles correspondent à la part '
        'estimée du secteur public hors CNSSAP, calculée par différence à partir des estimations '
        'globales de la Fonction publique : '
        '<span class="val" data-val-id="sB7-p1-d1" data-val-status="à valider" '
        'data-val-file="04_annexes/annexe_B_fiches_institutionnelles.md">1 622 972</span> '
        'pour chacune des années 2019 à 2022, '
        '<span class="val" data-val-id="sB7-p1-d2" data-val-status="à valider" '
        'data-val-file="04_annexes/annexe_B_fiches_institutionnelles.md">1 425 000</span> '
        'en 2023 et '
        '<span class="val" data-val-id="sB7-p1-d3" data-val-status="à valider" '
        'data-val-file="04_annexes/annexe_B_fiches_institutionnelles.md">1 727 000</span> '
        'en 2024. Source : estimations de la Fonction publique ; niveau de fiabilité : '
        'estimation provisoire, à confirmer par des données administratives détaillées.</span>\n'
        if institution == "TRESOR" else ""
    )
    contributors_term = (
        "personnes potentiellement couvertes estimées"
        if institution == "TRESOR" else "cotisants"
    )
    regimes_heading = "Dispositif représenté" if institution == "TRESOR" else "Régimes gérés"
    sex_heading = (
        "Répartition par sexe (personnes potentiellement couvertes estimées et bénéficiaires cumulés)"
        if institution == "TRESOR"
        else "Répartition par sexe (cotisants et bénéficiaires cumulés)"
    )
    settings_note = (
        '<p class="dev-note">Les règles enregistrées dans le formulaire institutionnel '
        "(Q1, Q1b, Q2 et Q4) sont appliquées aux visuels. Les données détaillées restent "
        "présentées par régime, sans déduplication.</p>\n"
        if settings else
        '<p class="dev-note">Aucun paramètre institutionnel enregistré : les visuels '
        "présentent les données ESS sans correction issue du formulaire.</p>\n"
    )

    if section_no:
        regimes_caption = f'<p class="table-caption"><strong>Tableau B.{section_no}.{table_no}</strong> — Régimes gérés, {label}</p>\n\n'
        table_no += 1
        charts_caption = (
            f'<p class="fig-caption"><strong>Figure B.{section_no}.{figure_no}</strong> — '
            f"Évolution des {contributors_term}, bénéficiaires, dépenses et recettes (tous régimes), {label} "
            f"(2019–2025)</p>\n\n"
        )
        figure_no += 1
        sex_caption = (
            f'<p class="fig-caption"><strong>Figure B.{section_no}.{figure_no}</strong> — '
            f"Répartition par sexe des {contributors_term} et bénéficiaires cumulés, {label} (2019–2025)</p>\n\n"
        )
        figure_no += 1
        detail_caption = f'<p class="table-caption"><strong>Tableau B.{section_no}.{table_no}</strong> — Données détaillées par régime et année, {label} (2019–2025){family_beneficiaries_footnote}</p>\n\n'
    else:
        regimes_caption = charts_caption = sex_caption = detail_caption = ""

    return (
        f"\n### {regimes_heading}\n\n"
        f"{regimes_caption}{regimes_md}\n"
        "### Aperçu graphique (tous régimes, toutes années)\n\n"
        f"{charts_caption}{charts_md}\n"
        f"### {sex_heading}\n\n"
        f"{sex_caption}{sex_md}\n"
        "### Données détaillées (par régime et année)\n\n"
        f"{detail_caption}{detail_md}{tresor_detail_footnote}\n"
        "*Source : base consolidée des ESS OIT/BIT.*\n"
        f"{settings_note}"
        '<p class="dev-note">Visuels et tableaux générés automatiquement, sans navigateur, '
        "via `py 09_scripts/generer_annexe_b_visuels.py`.</p>\n"
    )


def inject_into_markdown(md_text: str, institution: str, content: str) -> tuple[str, bool]:
    """Remplace uniquement le bloc délimité par les marqueurs AUTO_GENERE de l'institution."""
    start_marker = f"<!-- AUTO_GENERE:{institution}:DEBUT -->"
    end_marker = f"<!-- AUTO_GENERE:{institution}:FIN -->"
    pattern = re.compile(re.escape(start_marker) + r".*?" + re.escape(end_marker), re.DOTALL)
    if not pattern.search(md_text):
        return md_text, False
    replacement = f"{start_marker}\n{content}\n{end_marker}"
    # Callback (pas une chaîne) : re.sub n'interprète alors aucune séquence d'échappement.
    return pattern.sub(lambda _m: replacement, md_text, count=1), True


def has_auto_generated_block(md_text: str, institution: str) -> bool:
    """Indique si l'institution possède un bloc graphique dans l'annexe B."""
    start_marker = f"<!-- AUTO_GENERE:{institution}:DEBUT -->"
    end_marker = f"<!-- AUTO_GENERE:{institution}:FIN -->"
    return start_marker in md_text and end_marker in md_text


def main() -> int:
    if not DB_PATH.exists():
        print(f"  Base introuvable : {DB_PATH}")
        print("  Lancer d'abord : py 09_scripts/extraire_ess.py")
        return 1
    if not ANNEXE_B_MD.exists():
        print(f"  Fichier introuvable : {ANNEXE_B_MD}")
        return 1

    # Démarrer un serveur kaleido persistant : réutilisé pour tous les exports
    # d'images de ce lot (~80 images), au lieu de relancer un navigateur headless
    # à chaque appel — beaucoup plus rapide et beaucoup plus stable.
    server_started = False
    if kaleido is not None:
        try:
            kaleido.start_sync_server(silence_warnings=True)
            server_started = True
        except Exception as exc:
            print(f"  [AVERT] Démarrage du serveur kaleido persistant impossible ({exc}) — "
                  "poursuite en mode ponctuel (plus lent).")

    print("  Lecture BDD…")
    regimes, _prestations, regime_meta, _prestation_meta = load_all(DB_PATH)
    try:
        questionnaire_data = load_questionnaire_data()
    except (FileNotFoundError, ValueError) as exc:
        print(f"  Paramètres du formulaire invalides : {exc}")
        return 1
    institutions = sorted(set(r["institution"] for r in regimes))
    print(f"  {len(institutions)} institution(s) détectée(s) : {', '.join(institutions)}")

    md_text = ANNEXE_B_MD.read_text(encoding="utf-8-sig")
    total_updated = 0
    total_skipped = []
    all_chart_payloads: list[dict] = []

    for inst in institutions:
        print(f"  -> {inst}")
        if not has_auto_generated_block(md_text, inst):
            total_skipped.append(inst)
            continue
        content = build_section_content(
            regimes,
            regime_meta,
            inst,
            md_text,
            questionnaire_data.get(inst) or {},
            all_chart_payloads,
        )
        md_text, updated = inject_into_markdown(md_text, inst, content)
        if updated:
            total_updated += 1
        else:
            total_skipped.append(inst)

    ANNEXE_B_MD.write_text(md_text, encoding="utf-8-sig")
    export_annexe_b_word_data(all_chart_payloads)

    if server_started:
        try:
            kaleido.stop_sync_server()
        except Exception:
            pass

    print(f"\n  {total_updated} section(s) mise(s) à jour dans {ANNEXE_B_MD.name}")
    if total_skipped:
        print(f"  {len(total_skipped)} institution(s) sans marqueur AUTO_GENERE (ignorée(s)) : "
              f"{', '.join(total_skipped)}")
    return 0


if __name__ == "__main__":
    _exit_code = main()
    # Sortie forcée : kaleido/choreographer peut bloquer indéfiniment lors du
    # nettoyage normal de l'interpréteur (processus navigateur headless résiduel).
    # Tout le travail (images, fichier Markdown) est déjà écrit sur disque à ce stade.
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(_exit_code)
