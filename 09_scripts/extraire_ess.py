"""
extraire_ess.py — Extracteur spécifique aux fichiers ESS (format OIT/BIT)
=========================================================================
Lit les fichiers .xlsm "Enquête sur les Sources Statistiques" et peuple
la base protection_sociale_rdc.db via db_schema.py.

Ce script est propre au format ESS. D'autres scripts peuvent coexister pour
d'autres sources (rapport_annuel_cnss.py, publications_fss.py, etc.) en
utilisant les mêmes fonctions d'ingestion de db_schema.py.

Usage :
    py extraire_ess.py                    # traite tous les fichiers ESS présents dans 06_sources/ESS
    py extraire_ess.py --annee 2022       # seulement l'année 2022
    py extraire_ess.py --institution CNSS # seulement la CNSS
    py extraire_ess.py --inbox            # traite les fichiers déposés dans 06_sources/_entrants/
    py validate_ess.py                   # génère un rapport de validation sans écrire en base
    py extraire_ess.py --delete --source-id 42 --dry-run
    py extraire_ess.py --dry-run          # simule sans écrire dans la BDD
    py extraire_ess.py --verbose          # affiche le détail de chaque ligne

Structure du format ESS (gabarit OIT francophone) :
  Feuille "Inventaire des régimes" :
    - L1 : Pays, Contact
    - L2 : Période, E-mail
    - L4 : En-têtes colonnes (fonctions OIT, cotisants, bénéficiaires...)
    - L5 : Sous-en-têtes (Total/H/F, unités...)
    - L6+ : Un régime par ligne
  Feuilles de prestations ("Prestations aux familles", "Risques professionnels"...) :
    - L1 : Nom du régime, Année
    - L4 : En-têtes colonnes
    - L5+ : Une prestation par ligne (max 15)
    - L21 (env.) : Notes et Sources
"""

import os
import sys
import json
import argparse
import re
import shutil
import time
import warnings
import unicodedata
import sqlite3
from datetime import datetime
from pathlib import Path
from difflib import get_close_matches
warnings.filterwarnings('ignore')  # Supprime les warnings openpyxl sur extensions VBA

import openpyxl
from openpyxl.utils import get_column_letter

# Importation du module partagé (dans le même dossier)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db_schema import (get_db, create_or_update_db, register_source,
                       upsert_regime, upsert_indicateurs, upsert_prestation,
                       to_float, to_str, DB_PATH as SQLITE_DB_PATH)

# ---------------------------------------------------------------------------
# Répertoire de base des fichiers ESS
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ESS_BASE_DIR = os.path.normpath(os.path.join(_SCRIPT_DIR, '..', '06_sources', 'ESS'))
ESS_INBOX_DIR = os.path.normpath(os.path.join(_SCRIPT_DIR, '..', '06_sources', '_entrants'))
ESS_REPORT_DIR = os.path.normpath(os.path.join(_SCRIPT_DIR, '..', '10_output', 'ess_reports'))
ESS_DEST_DIRS = {
    'CNSS': os.path.normpath(os.path.join(ESS_BASE_DIR, 'ESS_CNSS')),
    'CNSSAP': os.path.normpath(os.path.join(ESS_BASE_DIR, 'ESS_CNSSAP')),
    'RDC': os.path.normpath(os.path.join(ESS_BASE_DIR, 'ESS_RDC_tous_regimes')),
}

# ---------------------------------------------------------------------------
# Notes connues associées à certains fichiers ESS archivés
# ---------------------------------------------------------------------------
ESS_SOURCE_NOTES = {
    os.path.normpath("ESS_CNSSAP/ESS CNSSAP 2020.xlsm"):
        "Date interne inventaire affiche 2022 — anomalie connue ; données retenues pour 2020",
    os.path.normpath("ESS_CNSSAP/ESS CNSSAP 2021.xlsm"):
        "Date interne inventaire affiche 2022 — anomalie connue ; données retenues pour 2021",
}

# Fonctions OIT dans l'ordre des colonnes du gabarit ESS (colonnes 4 à 20 de l'inventaire)
# Source : gabarit OIT/BIT standard francophone
OIT_FONCTIONS = [
    "Vieillesse",
    "Invalidité / Handicap",
    "Survivances",
    "Maladie (en espèces)",
    "Maternité / Paternité",
    "Enfants",
    "Famille",
    "Accès à l'éducation",
    "Chômage",
    "Accident du travail",
    "Soins de santé",
    "Services d'assistance sociale",
    "Programmes de travaux publics",
    "Subventions au logement",
    "Alimentation / Nutrition",
    "Assistance sociale ciblée pour la réduction de la pauvreté",
    "Autre soutien et assistance n.c.a.",
]
OIT_FONCTIONS_START_COL = 4   # index 0-basé de la première colonne de fonction OIT

# Colonnes de l'inventaire (indices 0-basés, vérifiés empiriquement)
INV_COL = {
    'regime_label':      0,   # "Régime 1", "Régime 2"...
    'nom_original':      1,
    'nom_fr':            2,
    'administrateur':    3,
    # fonctions_oit :    4 à 20 (17 colonnes)
    'type_financement':  21,
    'caractere':         22,
    'cotisants_total':   23,
    'cotisants_h':       24,
    'cotisants_f':       25,
    'beneficiaires_total': 26,
    'beneficiaires_h':   27,
    'beneficiaires_f':   28,
    'gestion_admin':     29,
    'gestion_op':        30,
    'type_assurance':    31,
    # Finances (inventaire ESS) :
    # 32 = total des dépenses, 33 = dépenses administratives, 34 = total des recettes
    'depenses_cdf':      32,
    'depenses_admin_cdf': 33,
    'recettes_cdf':      34,
    # Colonne USD non utilisée ici (le gabarit ESS observé est en monnaie locale)
    'recettes_usd':      None,
}

# Colonnes des feuilles de prestation (indices 0-basés, vérifiés empiriquement)
PREST_COL = {
    'prestation_label':      0,   # "Prestation 1"...
    'nom_original':          1,
    'nom_fr':                2,
    'fonction_oit':          3,
    'groupe_population':     4,
    'groupe_age':            5,
    'zone_geo':              6,
    'type_financement':      7,
    'couverture_total':      8,
    'couverture_h':          9,
    'couverture_f':         10,
    'beneficiaires_total':  11,
    'beneficiaires_h':      12,
    'beneficiaires_f':      13,
    'type_paiement':        14,
    'periodicite':          15,
    'montant_unitaire_cdf': 16,   # parfois vide pour CNSSAP (utilise col 17)
    'montant_unitaire_usd': 17,
    'critere_eligibilite':  18,
    'duree_service':        19,
    'col_20':               20,
    'col_21':               21,
    'age_legal_h':          22,
    'age_legal_f':          23,
    'condition_compl':      24,   # ex: 'Carrière' (CNSSAP)
    'depenses_regime_cdf':  25,   # dépenses totales du régime (répétées sur chaque ligne)
}

# ---------------------------------------------------------------------------
# Correspondance nom de feuille → code régime stable
# ---------------------------------------------------------------------------
SHEET_TO_REGIME = {
    'Prestations aux familles': 'R1',
    'Risques professionnels':   'R2',
    'Pension':                  'R3',
    'Régime 4':                 'R4',
    'CNSAP Régime de base':     'R1',   # CNSSAP
    'Reforme du transfert':     'R2',   # CNSSAP 2022
}


# ---------------------------------------------------------------------------
# Détection et routage des fichiers ESS entrants
# ---------------------------------------------------------------------------

_ESS_EXTENSION_RE = re.compile(r'\.(xlsx|xlsm|xls)$', re.IGNORECASE)
_ESS_YEAR_RE = re.compile(r'(19|20)\d{2}')
_ESS_INSTITUTION_RE = re.compile(r'\bESS[\s_-]*([A-Za-z0-9]{2,32})\b', re.IGNORECASE)
_REGIME_SHEET_RE = re.compile(r'^(?:regime|régime)\s*(\d+)\s*$', re.IGNORECASE)
_GENERIC_TOKENS = {
    "RDC", "REPUBLIQUE", "DEMOCRATIQUE", "CONGO", "PAYS", "CONTACT", "PERIODE",
    "DIRECTION", "ETUDES", "ORGANISATION", "MINISTERE", "MINISTERES",
    "ENSEIGNEMENT", "NOM", "REGIME", "PROGRAMME", "ASSURANCE", "SOCIALE",
    "SERVICES", "PUBLICS", "ETAT", "FOND", "SOLIDARITE", "SANTE", "AN",
}


def _safe_year(value):
    """Retourne une année entière si la valeur peut être interprétée comme telle."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    match = _ESS_YEAR_RE.search(str(value))
    if match:
        return int(match.group(0))
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _infer_ess_institution(name, workbook=None):
    """Déduit l'institution ESS à partir du nom du fichier puis, si besoin, du classeur."""
    if not name:
        return None

    base = os.path.splitext(os.path.basename(name))[0]
    upper = unicodedata.normalize("NFKD", base).encode("ascii", "ignore").decode("ascii").upper()
    if "CNSSAP" in upper:
        return "CNSSAP"
    if re.search(r"\bCNSS\b", upper):
        return "CNSS"
    if "RDC" in upper or "TOUS REGIMES" in upper:
        return "RDC"

    match = _ESS_INSTITUTION_RE.search(upper)
    if match:
        token = _sanitize_institution_code(match.group(1))
        if token:
            return token

    if workbook is not None:
        from_workbook = _infer_ess_institution_from_workbook(workbook)
        if from_workbook:
            return from_workbook
        sheetnames = set(workbook.sheetnames)
        if 'CNSAP Régime de base' in sheetnames or 'Reforme du transfert' in sheetnames:
            return 'CNSSAP'
        if {'Prestations aux familles', 'Risques professionnels', 'Pension'} & sheetnames:
            return 'CNSS'

    return None


def _infer_ess_institution_from_path(filepath):
    """Déduit l'institution ESS à partir du chemin d'archive."""
    if not filepath:
        return None
    normalized = os.path.normpath(os.path.dirname(filepath))
    parts = [p for p in normalized.split(os.sep) if p]
    for part in reversed(parts):
        upper = unicodedata.normalize("NFKD", part).encode("ascii", "ignore").decode("ascii").upper()
        if upper.startswith("ESS_"):
            suffix = upper[4:]
            if suffix.startswith("RDC"):
                return "RDC"
            token = _sanitize_institution_code(suffix)
            if token:
                return token
    return None


def _sanitize_institution_code(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text or None


def _extract_institution_from_text(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    upper = text.upper().strip()
    if not upper:
        return None
    if "CNSSAP" in upper:
        return "CNSSAP"
    if re.search(r"\bCNSS\b", upper):
        return "CNSS"
    if "RDC" in upper and ("TOUS REGIMES" in upper or "REPUBLIQUE DEMOCRATIQUE DU CONGO" in upper):
        return "RDC"

    for token in re.findall(r"\(([A-Z0-9]{2,32})\)", upper):
        token = _sanitize_institution_code(token)
        if token and token not in _GENERIC_TOKENS:
            return token

    for pattern in (r"^([A-Z0-9]{2,32})\s*[-:/]", r"^([A-Z0-9]{2,32})\b"):
        m = re.search(pattern, upper)
        if m:
            token = _sanitize_institution_code(m.group(1))
            if token and token not in _GENERIC_TOKENS:
                return token
    return None


def _infer_ess_institution_from_workbook(workbook):
    sheet_name = "Inventaire des régimes"
    if sheet_name not in workbook.sheetnames:
        return None
    ws = workbook[sheet_name]

    header_candidates = [ws["B1"].value, ws["D1"].value]
    for value in header_candidates:
        token = _extract_institution_from_text(value)
        if token:
            return token

    try:
        for row in ws.iter_rows(min_row=6, max_row=60, values_only=True):
            if not _is_regime_row(row):
                continue
            for idx in (INV_COL["administrateur"], INV_COL["nom_fr"], INV_COL["nom_original"]):
                token = _extract_institution_from_text(_get_cell(row, idx))
                if token:
                    return token
    except Exception:
        return None
    return None


def _infer_ess_institution_for_file(filepath, workbook=None):
    """Déduit l'institution depuis le nom, le chemin, puis le contenu si besoin."""
    return (
        _infer_ess_institution(os.path.basename(filepath), workbook)
        or _infer_ess_institution_from_path(filepath)
    )


def _infer_ess_year(filepath, ws=None):
    """Déduit l'année ESS à partir du nom du fichier puis de la feuille inventaire."""
    base = os.path.basename(filepath)
    year = _safe_year(base)
    if year:
        return year

    if ws is not None and 'Inventaire des régimes' in ws.parent.sheetnames:
        try:
            val = ws['B2'].value
            year = _safe_year(val)
            if year:
                return year
        except Exception:
            pass

    return None


def _collect_ess_year_candidates(filepath, workbook=None):
    """Collecte les années visibles dans le nom de fichier, l'inventaire et les onglets."""
    candidates = {
        "filename_year": _safe_year(os.path.basename(filepath)),
        "inventory_year": None,
        "sheet_years": {},
    }
    if workbook is None:
        return candidates

    if "Inventaire des régimes" in workbook.sheetnames:
        try:
            candidates["inventory_year"] = _safe_year(workbook["Inventaire des régimes"]["B2"].value)
        except Exception:
            candidates["inventory_year"] = None

    for sheet_name in workbook.sheetnames:
        if sheet_name in {"INSTRUCTIONS", "Inventaire des régimes", "Info Pauvreté", "HID_dropdown", "ADDITIONAL statistics for SDGs", "CALCULATIONS"}:
            continue
        try:
            year = _safe_year(workbook[sheet_name]["D1"].value)
        except Exception:
            year = None
        if year is not None:
            candidates["sheet_years"][sheet_name] = year

    return candidates


def _resolve_ess_year_consistency(filepath, workbook=None):
    """Résout l'année ESS si toutes les sources connues convergent, sinon retourne une incohérence."""
    candidates = _collect_ess_year_candidates(filepath, workbook)
    values = []
    for key in ("filename_year", "inventory_year"):
        if candidates.get(key) is not None:
            values.append(candidates[key])
    for value in candidates.get("sheet_years", {}).values():
        if value is not None:
            values.append(value)
    unique_values = sorted(set(values))
    if not unique_values:
        return {"year": None, "candidates": candidates, "issues": ["Aucune année détectable."]}
    if len(unique_values) > 1:
        return {
            "year": None,
            "candidates": candidates,
            "issues": [f"Années incohérentes détectées: {', '.join(str(v) for v in unique_values)}"],
        }
    return {"year": unique_values[0], "candidates": candidates, "issues": []}


def _select_ess_import_year(candidates, fallback=None):
    """Choisit l'année à utiliser pour l'import à partir des candidats visibles."""
    if fallback is not None:
        return fallback
    if candidates.get("filename_year") is not None:
        return candidates["filename_year"]
    if candidates.get("inventory_year") is not None:
        return candidates["inventory_year"]
    sheet_years = candidates.get("sheet_years") or {}
    for _sheet_name, year in sheet_years.items():
        if year is not None:
            return year
    return None


def _ess_destination_dir(institution):
    if not institution:
        return None
    normalized = _sanitize_institution_code(institution)
    if not normalized:
        return None
    if normalized in ESS_DEST_DIRS:
        return ESS_DEST_DIRS[normalized]
    if normalized == "RDC":
        return ESS_DEST_DIRS["RDC"]
    return os.path.normpath(os.path.join(ESS_BASE_DIR, f"ESS_{normalized}"))


def _ess_destination_name(filepath, institution, annee):
    """Normalise le nom cible d'un fichier ESS entrant."""
    base = os.path.basename(filepath)
    root, ext = os.path.splitext(base)
    lowered = root.lower()

    inst = _sanitize_institution_code(institution) if institution else None

    if inst == 'RDC' and re.search(r'^(ess[\s_-]+)?(rdc|ess rdc tous regimes)', lowered):
        return base

    if inst and re.search(rf'^ess[\s_-]+{inst.lower()}(?:[\s_-]+\d{{4}})?', lowered):
        return base

    if inst == 'RDC':
        suffix = f" {annee}" if annee else ""
        return f"ESS RDC tous regimes{suffix}{ext}"

    if inst:
        suffix = f" {annee}" if annee else ""
        return f"ESS {inst}{suffix}{ext}"

    return base


def _resolve_sheet_regime_suffix(sheet_name):
    if sheet_name in SHEET_TO_REGIME:
        return SHEET_TO_REGIME[sheet_name]
    match = _REGIME_SHEET_RE.match(str(sheet_name or "").strip())
    if match:
        return f"R{int(match.group(1))}"
    return None


def _is_ess_excel_candidate(filename):
    if not filename or filename.lower() == 'readme.md' or filename.startswith('~$'):
        return False
    return bool(_ESS_EXTENSION_RE.search(filename))


def _is_noncanonical_ess_copy(filename):
    lowered = os.path.splitext(os.path.basename(filename or ""))[0].lower()
    markers = (" copy", "_copy", "- copy", "copie", "anomal", "backup", "sauvegarde", "test")
    return any(marker in lowered for marker in markers)


def _build_ess_archive_entry(filepath, root_dir):
    rel_path = os.path.normpath(os.path.relpath(filepath, start=root_dir))
    institution = _infer_ess_institution_for_file(filepath)
    annee = _safe_year(os.path.basename(filepath))
    expected_name = _ess_destination_name(filepath, institution, annee) if institution and annee else os.path.basename(filepath)
    entry = {
        "filepath": filepath,
        "rel_path": rel_path,
        "institution": institution,
        "annee": annee,
        "note_anomalie": ESS_SOURCE_NOTES.get(rel_path),
        "is_canonical_name": os.path.normcase(os.path.basename(filepath)) == os.path.normcase(expected_name),
        "looks_like_copy": _is_noncanonical_ess_copy(filepath),
    }
    return entry


def _sort_ess_archive_entries(entries):
    return sorted(
        entries,
        key=lambda entry: (
            entry.get("institution") or "",
            entry.get("annee") or 0,
            entry.get("rel_path") or "",
        ),
    )


def discover_ess_archive_files(root_dir=ESS_BASE_DIR):
    """Scanne l'archive ESS et retient au plus un fichier par couple institution/année."""
    if not os.path.isdir(root_dir):
        return [], []

    raw_entries = []
    for root, _dirs, files in os.walk(root_dir):
        for filename in files:
            if not _is_ess_excel_candidate(filename):
                continue
            filepath = os.path.join(root, filename)
            raw_entries.append(_build_ess_archive_entry(filepath, root_dir))

    grouped = {}
    passthrough = []
    for entry in raw_entries:
        key = (entry.get("institution"), entry.get("annee"))
        if key[0] and key[1]:
            grouped.setdefault(key, []).append(entry)
        else:
            passthrough.append(entry)

    selected = []
    skipped = []

    for key, entries in grouped.items():
        ranked = sorted(
            entries,
            key=lambda entry: (
                0 if entry.get("is_canonical_name") else 1,
                1 if entry.get("looks_like_copy") else 0,
                len(entry.get("rel_path") or ""),
                entry.get("rel_path") or "",
            ),
        )
        selected.append(ranked[0])
        for duplicate in ranked[1:]:
            duplicate = dict(duplicate)
            duplicate["skip_reason"] = (
                f"Doublon archive pour {key[0]} {key[1]} ; fichier retenu : {ranked[0]['rel_path']}"
            )
            skipped.append(duplicate)

    selected.extend(passthrough)
    return _sort_ess_archive_entries(selected), _sort_ess_archive_entries(skipped)


def discover_ess_inbox_files(inbox_dir):
    """Retourne les fichiers ESS déposés dans le dossier de réception unique."""
    if not os.path.isdir(inbox_dir):
        return []

    candidates = []
    for root, _dirs, files in os.walk(inbox_dir):
        for filename in files:
            if not _is_ess_excel_candidate(filename):
                continue
            filepath = os.path.join(root, filename)
            candidates.append(filepath)

    return sorted(candidates)


def _move_file_with_retry(src_path, dst_path, attempts=6, delay_seconds=1.0):
    """
    Déplace un fichier avec plusieurs tentatives pour absorber les verrous
    temporaires (Excel, antivirus, synchronisation OneDrive).
    """
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            shutil.move(src_path, dst_path)
            return True, None
        except PermissionError as exc:
            last_error = exc
            if attempt < attempts:
                time.sleep(delay_seconds * attempt)
                continue
            break
        except OSError as exc:
            return False, str(exc)
    return False, str(last_error) if last_error else "Erreur de déplacement inconnue"


def _sanitize_slug(value):
    text = os.path.splitext(os.path.basename(value or ""))[0]
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("._-")
    return text or "ess"


def _normalize_label(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"^branche\s+des?\s+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _issue(severity, code, location, message, details=None):
    return {
        "severity": severity,
        "code": code,
        "location": location or "n/a",
        "message": message,
        "details": details or "",
    }


def _report_verdict(issues):
    severities = {item["severity"] for item in issues}
    if "error" in severities:
        return "blocked"
    if "warning" in severities:
        return "warnings"
    return "clean"


def _format_validation_report(report):
    lines = []
    issues = report.get("issues", [])
    counts = report.get("counts", {})
    checks = [
        "File identity and accessibility",
        "Workbook loadability",
        "Detected institution and year",
        "Year consistency across filename, inventory and sheets",
        "Archive duplicate file detection",
        "Database source duplicate detection",
        "Inventory sheet presence",
        "Regime row detection and numbering",
        "Regime sheet title vs inventory match",
        "Prestation row detection and numbering",
        "Numeric cell parsing on key fields",
        "Currency fallback risk detection",
        "Cross-sheet regime consistency",
    ]
    lines.append("# ESS validation report")
    lines.append("")
    lines.append(f"- Generated at: {report.get('generated_at', 'n/a')}")
    lines.append(f"- File: {report.get('file_name', 'n/a')}")
    lines.append(f"- Source path: {report.get('file_path', 'n/a')}")
    lines.append(f"- Mode: {report.get('mode', 'validation')}")
    lines.append(f"- Detected institution: {report.get('detected_institution', 'n/a')}")
    lines.append(f"- Detected year: {report.get('detected_year', 'n/a')}")
    lines.append(f"- Verdict: {report.get('verdict', 'n/a')}")
    lines.append("")
    lines.append("## Summary")
    lines.append(f"- Errors: {counts.get('errors', 0)}")
    lines.append(f"- Warnings: {counts.get('warnings', 0)}")
    lines.append(f"- Infos: {counts.get('infos', 0)}")
    lines.append(f"- Regimes found: {counts.get('regimes_found', 0)}")
    lines.append(f"- Prestations found: {counts.get('prestations_found', 0)}")
    lines.append("")
    lines.append("## Sheets")
    for sheet in report.get("sheet_names", []):
        lines.append(f"- {sheet}")
    if not report.get("sheet_names"):
        lines.append("- n/a")
    lines.append("")
    lines.append("## Issues")
    if not issues:
        lines.append("- None")
    else:
        lines.append("| Severity | Code | Location | Message | Details |")
        lines.append("|---|---|---|---|---|")
        for item in issues:
            details = (item.get("details") or "").replace("|", "\\|").replace("\n", " ")
            message = (item.get("message") or "").replace("|", "\\|")
            location = (item.get("location") or "").replace("|", "\\|")
            lines.append(
                f"| {item.get('severity','')} | {item.get('code','')} | {location} | {message} | {details} |"
            )
    lines.append("")
    lines.append("## Checks performed")
    for check in checks:
        lines.append(f"- {check}")
    lines.append("")
    return "\n".join(lines)


def _write_validation_report(report, report_dir=ESS_REPORT_DIR):
    os.makedirs(report_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = _sanitize_slug(report.get("file_name"))
    out_path = os.path.join(report_dir, f"{stamp}_{slug}_validation.md")
    Path(out_path).write_text(_format_validation_report(report), encoding="utf-8")
    return out_path


def _find_existing_ess_sources(institution, annee, nom_fichier):
    if not institution or annee is None or not os.path.exists(SQLITE_DB_PATH):
        return []
    conn = sqlite3.connect(SQLITE_DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """SELECT source_id, institution, annee_donnees, nom_fichier, chemin_fichier, date_ingestion
               FROM sources_ingestion
               WHERE type_source = 'ESS'
                 AND UPPER(institution) = UPPER(?)
                 AND annee_donnees = ?
                 AND nom_fichier = ?
               ORDER BY date_ingestion DESC""",
            (institution, annee, nom_fichier),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _check_numeric_cell(issues, raw_value, parsed_value, location, field_name, severity="warning"):
    if raw_value is None:
        return
    if isinstance(raw_value, str) and not raw_value.strip():
        return
    if parsed_value is None:
        issues.append(_issue(
            severity,
            "non_numeric_value",
            location,
            f"Valeur non numérique pour {field_name}",
            f"Contenu lu: {raw_value!r}",
        ))


def validate_ess_workbook(filepath, institution=None, annee=None, note_anomalie=None, source_mode="archive_scan"):
    base = os.path.basename(filepath)
    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "mode": "validation-only",
        "file_name": base,
        "file_path": os.path.abspath(filepath),
        "expected_institution": institution,
        "expected_year": annee,
        "detected_institution": institution,
        "detected_year": annee,
        "note": note_anomalie,
        "source_mode": source_mode,
        "sheet_names": [],
        "issues": [],
        "counts": {},
        "verdict": "blocked",
    }

    issues = report["issues"]

    if not os.path.exists(filepath):
        issues.append(_issue("error", "file_missing", base, "Fichier introuvable"))
        report["counts"] = {"errors": 1, "warnings": 0, "infos": 0, "regimes_found": 0, "prestations_found": 0}
        report["verdict"] = _report_verdict(issues)
        return report

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=False, data_only=True)
    except Exception as exc:
        issues.append(_issue("error", "workbook_unreadable", base, "Impossible d'ouvrir le classeur", str(exc)))
        report["counts"] = {"errors": 1, "warnings": 0, "infos": 0, "regimes_found": 0, "prestations_found": 0}
        report["verdict"] = _report_verdict(issues)
        return report

    try:
        sheetnames = list(wb.sheetnames)
        report["sheet_names"] = sheetnames

        if institution is None:
            institution = _infer_ess_institution_for_file(filepath, wb)
        if annee is None:
            annee = _infer_ess_year(filepath, wb["Inventaire des régimes"] if "Inventaire des régimes" in sheetnames else None)
        report["detected_institution"] = institution
        report["detected_year"] = annee

        if report["expected_institution"] and institution and report["expected_institution"].upper() != institution.upper():
            issues.append(_issue(
                "error",
                "institution_mismatch",
                base,
                "Institution détectée différente de l'institution attendue",
                f"Attendue: {report['expected_institution']!r}, détectée: {institution!r}",
            ))
        if report["expected_year"] and annee and report["expected_year"] != annee:
            issues.append(_issue(
                "warning",
                "year_mismatch",
                base,
                "Année détectée différente de l'année attendue",
                f"Attendue: {report['expected_year']!r}, détectée: {annee!r}",
            ))

        if note_anomalie:
            issues.append(_issue("info", "source_note", base, "Note associée à la source", note_anomalie))

        destination_dir = _ess_destination_dir(institution) if institution else None
        destination_name = _ess_destination_name(filepath, institution, annee) if institution else base
        destination_path = os.path.normpath(os.path.join(destination_dir, destination_name)) if destination_dir else None
        if destination_path and os.path.exists(destination_path):
            same_file = os.path.abspath(destination_path) == os.path.abspath(filepath)
            severity = "info" if same_file else "warning"
            issues.append(_issue(
                severity,
                "archive_file_exists",
                destination_path,
                "Un fichier portant ce nom existe déjà dans l'archive ESS",
                "Même chemin" if same_file else "Le fichier serait remplacé ou ambigu à l'import.",
            ))

        existing_sources = _find_existing_ess_sources(institution, annee, destination_name)
        for row in existing_sources:
            issues.append(_issue(
                "warning",
                "db_source_exists",
                row.get("nom_fichier") or destination_name,
                "La même source ESS existe déjà en base",
                f"source_id={row.get('source_id')}, institution={row.get('institution')}, annee={row.get('annee_donnees')}, date_ingestion={row.get('date_ingestion')}",
            ))

        if "Inventaire des régimes" not in sheetnames:
            suggestion = get_close_matches("Inventaire des régimes", sheetnames, n=1, cutoff=0.6)
            detail = "Aucune feuille d'inventaire trouvée."
            if suggestion:
                detail += f" Feuille proche détectée: {suggestion[0]!r}."
            issues.append(_issue("error", "missing_inventory_sheet", "Inventaire des régimes", detail))
            report["counts"] = {
                "errors": sum(1 for item in issues if item["severity"] == "error"),
                "warnings": sum(1 for item in issues if item["severity"] == "warning"),
                "infos": sum(1 for item in issues if item["severity"] == "info"),
                "regimes_found": 0,
                "prestations_found": 0,
            }
            report["verdict"] = _report_verdict(issues)
            return report

        ws_inv = wb["Inventaire des régimes"]
        inventory_year = _safe_year(ws_inv["B2"].value)
        if inventory_year is not None:
            if annee is not None and inventory_year != annee:
                issues.append(_issue(
                    "warning",
                    "inventory_year_mismatch",
                    "Inventaire des régimes!B2",
                    "Année de l'inventaire différente de l'année détectée",
                    f"Inventaire: {inventory_year!r}, détectée: {annee!r}",
                ))
        else:
            issues.append(_issue("warning", "missing_inventory_year", "Inventaire des régimes!B2", "Année non lisible dans l'inventaire"))

        rows = list(ws_inv.iter_rows(values_only=True))
        regime_rows = []
        regime_codes = set()
        regime_nums = set()
        for idx, row in enumerate(rows, start=1):
            if not _is_regime_row(row):
                continue
            regime_rows.append((idx, row))
            num = _extract_regime_num(to_str(_get_cell(row, INV_COL["regime_label"])))
            if num is None:
                issues.append(_issue("error", "invalid_regime_number", f"Inventaire!A{idx}", "Numéro de régime illisible"))
            elif num in regime_nums:
                issues.append(_issue("warning", "duplicate_regime_number", f"Inventaire!A{idx}", f"Régime {num} déjà rencontré"))
            else:
                regime_nums.add(num)
            regime_code = f"{institution}_R{num}" if institution and num is not None else None
            if regime_code:
                regime_codes.add(regime_code)

            numeric_fields = [
                ("cotisants_total", INV_COL["cotisants_total"]),
                ("cotisants_h", INV_COL["cotisants_h"]),
                ("cotisants_f", INV_COL["cotisants_f"]),
                ("beneficiaires_total", INV_COL["beneficiaires_total"]),
                ("beneficiaires_h", INV_COL["beneficiaires_h"]),
                ("beneficiaires_f", INV_COL["beneficiaires_f"]),
                ("depenses_cdf", INV_COL["depenses_cdf"]),
                ("depenses_admin_cdf", INV_COL["depenses_admin_cdf"]),
                ("recettes_cdf", INV_COL["recettes_cdf"]),
            ]
            for field_name, col_idx in numeric_fields:
                raw = _get_cell(row, col_idx)
                parsed = to_float(raw)
                _check_numeric_cell(
                    issues,
                    raw,
                    parsed,
                    f"Inventaire!{get_column_letter(col_idx + 1)}{idx}",
                    field_name,
                    severity="warning",
                )

        if not regime_rows:
            issues.append(_issue("error", "no_regime_rows", "Inventaire des régimes", "Aucune ligne de régime valide détectée"))

        inventory_labels = {}
        for idx, row in regime_rows:
            num = _extract_regime_num(to_str(_get_cell(row, INV_COL["regime_label"])))
            if num is None:
                continue
            expected_name = to_str(_get_cell(row, INV_COL["nom_fr"])) or to_str(_get_cell(row, INV_COL["nom_original"]))
            if expected_name:
                inventory_labels[num] = expected_name

        excluded_sheets = {"INSTRUCTIONS", "Inventaire des régimes", "Info Pauvreté", "HID_dropdown", "ADDITIONAL statistics for SDGs", "CALCULATIONS"}
        data_sheet_names = []
        for name in sheetnames:
            if name in excluded_sheets:
                continue
            title_b1 = to_str(wb[name]["B1"].value)
            if title_b1 and _normalize_label(title_b1).startswith("nom du regime"):
                continue
            data_sheet_names.append(name)
        matched_inventory_nums = set()
        for sheet_name in data_sheet_names:
            ws = wb[sheet_name]
            title_b1 = to_str(ws["B1"].value)
            sheet_year = _safe_year(ws["D1"].value)
            if sheet_year is None:
                issues.append(_issue("warning", "missing_sheet_year", f"{sheet_name}!D1", "Année non lisible dans l'onglet de prestation"))
            else:
                if annee is not None and sheet_year != annee:
                    issues.append(_issue(
                        "warning",
                        "sheet_year_mismatch",
                        f"{sheet_name}!D1",
                        "Année de l'onglet différente de l'année détectée",
                        f"Onglet: {sheet_year!r}, détectée: {annee!r}",
                    ))
                if inventory_year is not None and sheet_year != inventory_year:
                    issues.append(_issue(
                        "warning",
                        "sheet_inventory_year_mismatch",
                        f"{sheet_name}!D1",
                        "Année de l'onglet différente de l'année de l'inventaire",
                        f"Onglet: {sheet_year!r}, inventaire: {inventory_year!r}",
                    ))
            norm_title = _normalize_label(title_b1)
            matched_num = None
            norm_inventory = {num: _normalize_label(expected_name) for num, expected_name in inventory_labels.items()}
            norm_values = list(norm_inventory.values())
            close = get_close_matches(norm_title, norm_values, n=1, cutoff=0.85) if norm_title else []
            if close:
                close_norm = close[0]
                for num, expected_norm in norm_inventory.items():
                    if expected_norm == close_norm:
                        matched_num = num
                        break
            else:
                for num, expected_name in inventory_labels.items():
                    if _normalize_label(expected_name) == norm_title:
                        matched_num = num
                        break
            if matched_num is None:
                suggestion = get_close_matches(title_b1 or sheet_name, list(inventory_labels.values()), n=1, cutoff=0.6)
                detail = f"Le titre B1 {title_b1!r} ne correspond à aucun intitulé d'inventaire."
                if suggestion:
                    detail += f" Intitulé proche: {suggestion[0]!r}."
                issues.append(_issue("warning", "sheet_title_mismatch", sheet_name, detail))
            else:
                matched_inventory_nums.add(matched_num)

        for num, expected_name in inventory_labels.items():
            if num not in matched_inventory_nums:
                issues.append(_issue(
                    "warning",
                    "missing_prestation_sheet",
                    f"Régime {num}",
                    f"Aucune feuille de prestations ne correspond à l'intitulé d'inventaire {expected_name!r}.",
                ))

        prestation_codes_seen = set()
        prestations_found = 0
        for sheet_name in data_sheet_names:
            ws = wb[sheet_name]
            title_b1 = to_str(ws["B1"].value)
            matched_num = None
            matched_name = None
            for num, expected_name in inventory_labels.items():
                if _normalize_label(expected_name) == _normalize_label(title_b1):
                    matched_num = num
                    matched_name = expected_name
                    break
            regime_suffix = f"R{matched_num}" if matched_num is not None else None
            sheet_rows = list(ws.iter_rows(values_only=True))
            sheet_prestations = 0
            for idx, row in enumerate(sheet_rows, start=1):
                label = to_str(_get_cell(row, PREST_COL["prestation_label"]))
                nom_fr = to_str(_get_cell(row, PREST_COL["nom_fr"]))
                nom_orig = to_str(_get_cell(row, PREST_COL["nom_original"]))
                if not label and not nom_fr and not nom_orig:
                    continue
                if label and label.lower().startswith("prestation au titre"):
                    continue
                if label and not label.startswith("Prestation "):
                    continue
                if not label:
                    issues.append(_issue("warning", "missing_prestation_label", f"{sheet_name}!A{idx}", "Ligne de prestation sans libellé attendu"))
                    continue
                try:
                    num = int(label.replace("Prestation", "").strip())
                except ValueError:
                    continue
                relevant_data = [
                    _get_cell(row, PREST_COL["fonction_oit"]),
                    _get_cell(row, PREST_COL["groupe_population"]),
                    _get_cell(row, PREST_COL["groupe_age"]),
                    _get_cell(row, PREST_COL["zone_geo"]),
                    _get_cell(row, PREST_COL["type_financement"]),
                    _get_cell(row, PREST_COL["couverture_total"]),
                    _get_cell(row, PREST_COL["couverture_h"]),
                    _get_cell(row, PREST_COL["couverture_f"]),
                    _get_cell(row, PREST_COL["beneficiaires_total"]),
                    _get_cell(row, PREST_COL["beneficiaires_h"]),
                    _get_cell(row, PREST_COL["beneficiaires_f"]),
                    _get_cell(row, PREST_COL["type_paiement"]),
                    _get_cell(row, PREST_COL["periodicite"]),
                    _get_cell(row, PREST_COL["montant_unitaire_cdf"]),
                    _get_cell(row, PREST_COL["montant_unitaire_usd"]),
                    _get_cell(row, PREST_COL["critere_eligibilite"]),
                    _get_cell(row, PREST_COL["duree_service"]),
                    _get_cell(row, PREST_COL["age_legal_h"]),
                    _get_cell(row, PREST_COL["age_legal_f"]),
                    _get_cell(row, PREST_COL["condition_compl"]),
                    _get_cell(row, PREST_COL["depenses_regime_cdf"]),
                ]
                if not any(value is not None and not (isinstance(value, str) and not value.strip()) for value in relevant_data):
                    continue
                sheet_prestations += 1
                prestations_found += 1
                if institution and regime_suffix:
                    prestation_codes_seen.add(f"{institution}_{regime_suffix}")

                for field_name, col_idx in [
                    ("couverture_total", PREST_COL["couverture_total"]),
                    ("couverture_h", PREST_COL["couverture_h"]),
                    ("couverture_f", PREST_COL["couverture_f"]),
                    ("beneficiaires_total", PREST_COL["beneficiaires_total"]),
                    ("beneficiaires_h", PREST_COL["beneficiaires_h"]),
                    ("beneficiaires_f", PREST_COL["beneficiaires_f"]),
                    ("montant_unitaire_cdf", PREST_COL["montant_unitaire_cdf"]),
                    ("montant_unitaire_usd", PREST_COL["montant_unitaire_usd"]),
                    ("depenses_regime_cdf", PREST_COL["depenses_regime_cdf"]),
                ]:
                    raw = _get_cell(row, col_idx)
                    parsed = to_float(raw)
                    _check_numeric_cell(
                        issues,
                        raw,
                        parsed,
                        f"{sheet_name}!{get_column_letter(col_idx + 1)}{idx}",
                        field_name,
                        severity="warning",
                    )

                if not nom_fr and not nom_orig:
                    issues.append(_issue("warning", "missing_prestation_name", f"{sheet_name}!A{idx}", "Nom de prestation manquant"))
                if _get_cell(row, PREST_COL["montant_unitaire_cdf"]) is None and _get_cell(row, PREST_COL["montant_unitaire_usd"]) is not None:
                    issues.append(_issue("warning", "currency_fallback_risk", f"{sheet_name}!Q{idx}", "Montant présent en colonne USD mais absent en CDF", "Le convertisseur actuel peut mal interpréter cette valeur."))

            if sheet_prestations == 0:
                issues.append(_issue("warning", "empty_prestation_sheet", sheet_name, "Feuille présente mais aucune prestation valide détectée"))

        counts = {
            "errors": sum(1 for item in issues if item["severity"] == "error"),
            "warnings": sum(1 for item in issues if item["severity"] == "warning"),
            "infos": sum(1 for item in issues if item["severity"] == "info"),
            "regimes_found": len(regime_rows),
            "prestations_found": prestations_found,
        }
        report["counts"] = counts
        report["verdict"] = _report_verdict(issues)
        return report
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Fonctions de parsing
# ---------------------------------------------------------------------------

def _get_cell(row, idx, default=None):
    """Retourne la valeur d'une cellule par index, ou default si hors limites."""
    try:
        v = row[idx]
        if v is None:
            return default
        if isinstance(v, str) and v.strip() == '':
            return default
        return v
    except IndexError:
        return default


def _is_regime_row(row):
    """Vérifie qu'une ligne de l'inventaire est bien un régime déclaré (pas vide/gabarit)."""
    label = to_str(_get_cell(row, INV_COL['regime_label']))
    if not label or not label.startswith('Régime '):
        return False
    nom = to_str(_get_cell(row, INV_COL['nom_original']))
    if not nom or nom.startswith('Nom du régime'):
        return False
    return True


def _extract_regime_num(label):
    """Extrait le numéro du régime depuis 'Régime 3' → 3."""
    try:
        return int(label.replace('Régime', '').strip())
    except (ValueError, AttributeError):
        return None


def _extract_fonctions_oit(row):
    """Retourne la liste des fonctions OIT cochées (valeur 'X' ou 'x') dans la ligne."""
    fonctions = []
    for i, nom in enumerate(OIT_FONCTIONS):
        col = OIT_FONCTIONS_START_COL + i
        val = to_str(_get_cell(row, col))
        if val and val.upper() == 'X':
            fonctions.append(nom)
    return fonctions


def _detect_unite_monetaire(ws):
    """
    Détecte l'unité monétaire utilisée dans l'inventaire.
    CNSSAP utilise 'Milliards CDF', CNSS utilise 'CDF'.
    Lit la cellule à la position (ligne 5, col 33) dans la feuille inventaire.
    """
    try:
        rows = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))
        if rows:
            val = to_str(_get_cell(rows[0], 32))
            if val and 'illiard' in val:
                return 'Milliards_CDF', 1e9
    except Exception:
        pass
    return 'CDF', 1.0


def parse_inventaire(ws, institution, annee, verbose=False):
    """
    Parse la feuille "Inventaire des régimes".
    Retourne une liste de dicts, un par régime déclaré.
    """
    unite_mon, facteur = _detect_unite_monetaire(ws)
    rows = list(ws.iter_rows(values_only=True))
    regimes = []

    for i, row in enumerate(rows):
        if not _is_regime_row(row):
            continue

        num = _extract_regime_num(to_str(_get_cell(row, 0)))
        if num is None:
            continue

        regime_code = f"{institution}_R{num}"
        fonctions = _extract_fonctions_oit(row)

        # Cotisants (nettoyage des espaces dans les nombres CNSSAP ex: "172 304")
        def parse_num(idx):
            return to_float(_get_cell(row, idx))

        depenses_raw = parse_num(INV_COL['depenses_cdf'])
        depenses_admin_raw = parse_num(INV_COL['depenses_admin_cdf'])
        recettes_raw = parse_num(INV_COL['recettes_cdf'])

        nom_original = to_str(_get_cell(row, INV_COL['nom_original']))
        nom_fr = to_str(_get_cell(row, INV_COL['nom_fr'])) or nom_original
        if institution == 'CNSS' and regime_code == 'CNSS_R4':
            nom_fr = "Action sociale et sanitaire"

        rec = {
            'institution':      institution,
            'regime_code':      regime_code,
            'regime_num':       num,
            'annee':            annee,
            'nom_original':     nom_original,
            'nom_fr':           nom_fr,
            'administrateur':   to_str(_get_cell(row, INV_COL['administrateur'])),
            'type_financement': to_str(_get_cell(row, INV_COL['type_financement'])),
            'caractere':        to_str(_get_cell(row, INV_COL['caractere'])),
            'type_assurance':   to_str(_get_cell(row, INV_COL['type_assurance'])),
            'gestion':          to_str(_get_cell(row, INV_COL['gestion_admin'])),
            'fonctions_oit':    json.dumps(fonctions, ensure_ascii=False),
            # Indicateurs quantitatifs
            'cotisants_total':  parse_num(INV_COL['cotisants_total']),
            'cotisants_h':      parse_num(INV_COL['cotisants_h']),
            'cotisants_f':      parse_num(INV_COL['cotisants_f']),
            'beneficiaires_total': parse_num(INV_COL['beneficiaires_total']),
            'beneficiaires_h':  parse_num(INV_COL['beneficiaires_h']),
            'beneficiaires_f':  parse_num(INV_COL['beneficiaires_f']),
            # Finances (converties en CDF si source en Milliards)
            'recettes_cdf':               recettes_raw * facteur if recettes_raw is not None else None,
            'recettes_usd':               parse_num(INV_COL['recettes_usd']) if INV_COL['recettes_usd'] is not None else None,
            # Le gabarit ESS fournit "Total des dépenses". Nous l'utilisons comme proxy
            # de dépense de prestation dans le schéma actuel.
            'depenses_prestations_cdf':   depenses_raw * facteur if depenses_raw is not None else None,
            'depenses_admin_cdf':         depenses_admin_raw * facteur if depenses_admin_raw is not None else None,
            'unite_monetaire_source': unite_mon,
        }

        if verbose:
            print(f"    Régime {num} ({regime_code}): {rec['nom_fr']} | "
                  f"cotisants={rec['cotisants_total']} | "
                  f"bénéf={rec['beneficiaires_total']} | "
                  f"fonctions={fonctions}")

        regimes.append(rec)

    return regimes


def purge_existing_ess_import(conn, institution, annee, nom_fichier):
    """Supprime un import ESS existant pour éviter les doublons et remplacer les valeurs."""
    source_rows = conn.execute(
        """SELECT source_id
           FROM sources_ingestion
           WHERE type_source = 'ESS'
             AND institution = ?
             AND annee_donnees = ?
             AND nom_fichier = ?""",
        (institution, annee, nom_fichier),
    ).fetchall()
    source_ids = [r[0] for r in source_rows]
    if not source_ids:
        return 0

    placeholders = ",".join(["?"] * len(source_ids))
    conn.execute(f"DELETE FROM prestations_historique WHERE source_id IN ({placeholders})", source_ids)
    conn.execute(f"DELETE FROM indicateurs_regime WHERE source_id IN ({placeholders})", source_ids)
    conn.execute(f"DELETE FROM regimes_historique WHERE source_id IN ({placeholders})", source_ids)
    conn.execute(f"DELETE FROM sources_ingestion WHERE source_id IN ({placeholders})", source_ids)
    conn.commit()
    return len(source_ids)


def find_ess_sources(conn, institution=None, annee=None, nom_fichier=None, source_id=None):
    """Retourne les sources ESS correspondant aux filtres."""
    where = ["type_source = 'ESS'"]
    params = []

    if source_id is not None:
        where.append("source_id = ?")
        params.append(source_id)
    if institution:
        where.append("UPPER(institution) = UPPER(?)")
        params.append(institution)
    if annee is not None:
        where.append("annee_donnees = ?")
        params.append(annee)
    if nom_fichier:
        where.append("nom_fichier = ?")
        params.append(nom_fichier)

    sql = f"""SELECT source_id, institution, annee_donnees, nom_fichier, chemin_fichier, date_ingestion
              FROM sources_ingestion
              WHERE {' AND '.join(where)}
              ORDER BY institution, annee_donnees, source_id"""
    return conn.execute(sql, params).fetchall()


def delete_ess_sources(conn, source_ids, dry_run=False):
    """Supprime proprement un ensemble de source_id ESS avec leurs données liées."""
    if not source_ids:
        return {'sources': 0, 'regimes': 0, 'indicateurs': 0, 'prestations': 0}

    placeholders = ",".join(["?"] * len(source_ids))
    counts = {
        'sources': conn.execute(
            f"SELECT COUNT(*) FROM sources_ingestion WHERE source_id IN ({placeholders})", source_ids
        ).fetchone()[0],
        'regimes': conn.execute(
            f"SELECT COUNT(*) FROM regimes_historique WHERE source_id IN ({placeholders})", source_ids
        ).fetchone()[0],
        'indicateurs': conn.execute(
            f"SELECT COUNT(*) FROM indicateurs_regime WHERE source_id IN ({placeholders})", source_ids
        ).fetchone()[0],
        'prestations': conn.execute(
            f"SELECT COUNT(*) FROM prestations_historique WHERE source_id IN ({placeholders})", source_ids
        ).fetchone()[0],
    }

    if dry_run:
        return counts

    conn.execute(f"DELETE FROM prestations_historique WHERE source_id IN ({placeholders})", source_ids)
    conn.execute(f"DELETE FROM indicateurs_regime WHERE source_id IN ({placeholders})", source_ids)
    conn.execute(f"DELETE FROM regimes_historique WHERE source_id IN ({placeholders})", source_ids)
    conn.execute(f"DELETE FROM sources_ingestion WHERE source_id IN ({placeholders})", source_ids)
    conn.commit()
    return counts


def parse_prestation_sheet(ws, institution, regime_code, annee, verbose=False):
    """
    Parse une feuille de prestation (ex: "Prestations aux familles").
    Retourne une liste de dicts, un par prestation déclarée (lignes non vides).
    """
    rows = list(ws.iter_rows(values_only=True))
    prestations = []

    for i, row in enumerate(rows):
        label = to_str(_get_cell(row, PREST_COL['prestation_label']))
        if not label or not label.startswith('Prestation '):
            continue

        num_str = label.replace('Prestation', '').strip()
        try:
            num = int(num_str)
        except ValueError:
            continue

        nom_fr = to_str(_get_cell(row, PREST_COL['nom_fr']))
        nom_orig = to_str(_get_cell(row, PREST_COL['nom_original']))

        # Ligne vide = fin des prestations déclarées
        if not nom_fr and not nom_orig:
            continue

        # Montant unitaire : col 16, fallback col 17 (certains fichiers ESS utilisent seulement la 2e colonne)
        montant_cdf = to_float(_get_cell(row, PREST_COL['montant_unitaire_cdf']))
        montant_cdf_alt = to_float(_get_cell(row, PREST_COL['montant_unitaire_usd']))
        if montant_cdf is None and montant_cdf_alt is not None:
            montant_cdf = montant_cdf_alt
        # Les fichiers ESS traités ici renseignent les montants en monnaie locale;
        # la colonne USD n'est donc pas alimentée à ce stade.
        montant_usd = None

        # Dépenses régime totales (répétées sur chaque ligne)
        dep_regime = to_float(_get_cell(row, PREST_COL['depenses_regime_cdf']))

        prest = {
            'institution':              institution,
            'regime_code':              regime_code,
            'annee':                    annee,
            'prestation_num':           num,
            'nom_original':             nom_orig,
            'nom_fr':                   nom_fr,
            'fonction_oit':             to_str(_get_cell(row, PREST_COL['fonction_oit'])),
            'groupe_population':        to_str(_get_cell(row, PREST_COL['groupe_population'])),
            'groupe_age':               to_str(_get_cell(row, PREST_COL['groupe_age'])),
            'zone_geo':                 to_str(_get_cell(row, PREST_COL['zone_geo'])),
            'type_financement':         to_str(_get_cell(row, PREST_COL['type_financement'])),
            'couverture_effective_total': to_float(_get_cell(row, PREST_COL['couverture_total'])),
            'couverture_h':             to_float(_get_cell(row, PREST_COL['couverture_h'])),
            'couverture_f':             to_float(_get_cell(row, PREST_COL['couverture_f'])),
            'beneficiaires_total':      to_float(_get_cell(row, PREST_COL['beneficiaires_total'])),
            'beneficiaires_h':          to_float(_get_cell(row, PREST_COL['beneficiaires_h'])),
            'beneficiaires_f':          to_float(_get_cell(row, PREST_COL['beneficiaires_f'])),
            'type_paiement':            to_str(_get_cell(row, PREST_COL['type_paiement'])),
            'periodicite':              to_str(_get_cell(row, PREST_COL['periodicite'])),
            'montant_unitaire_cdf':     montant_cdf,
            'montant_unitaire_usd':     montant_usd,
            'critere_eligibilite':      to_str(_get_cell(row, PREST_COL['critere_eligibilite'])),
            'duree_service_requise':    to_str(_get_cell(row, PREST_COL['duree_service'])),
            'age_legal_h':              to_str(_get_cell(row, PREST_COL['age_legal_h'])),
            'age_legal_f':              to_str(_get_cell(row, PREST_COL['age_legal_f'])),
            'condition_complementaire': to_str(_get_cell(row, PREST_COL['condition_compl'])),
            'depenses_regime_cdf':      dep_regime,
        }

        if verbose:
            print(f"      Prestation {num}: {nom_fr} | "
                  f"bénéf={prest['beneficiaires_total']} | "
                  f"montant={montant_cdf} CDF")

        prestations.append(prest)

    return prestations


# ---------------------------------------------------------------------------
# Traitement d'un fichier ESS
# ---------------------------------------------------------------------------

def process_ess_file(filepath, institution, annee, note_anomalie=None,
                     conn=None, dry_run=False, verbose=False,
                     nom_fichier_override=None, chemin_rel_override=None):
    """
    Traite un fichier ESS complet :
    1. Enregistre la source dans sources_ingestion
    2. Parse l'inventaire → regimes_historique + indicateurs_regime
    3. Parse chaque feuille de prestation → prestations_historique
    """
    institution = _sanitize_institution_code(institution)
    nom_fichier = nom_fichier_override or os.path.basename(filepath)
    chemin_rel  = chemin_rel_override or os.path.relpath(filepath, start=os.path.normpath(
                      os.path.join(_SCRIPT_DIR, '..')))

    if not institution:
        print(f"  ✗ Institution ESS introuvable pour {nom_fichier}")
        return False

    print(f"\n{'─'*65}")
    print(f"  {institution} {annee}  —  {nom_fichier}")
    if note_anomalie:
        print(f"  ⚠ {note_anomalie}")

    if not os.path.exists(filepath):
        print(f"  ✗ Fichier introuvable : {filepath}")
        return False

    # Charger le classeur (data_only=True : lit les valeurs calculées, pas les formules)
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=False, data_only=True)
    print(f"  Feuilles : {wb.sheetnames}")

    year_check = _resolve_ess_year_consistency(filepath, wb)
    year_candidates = year_check.get("candidates") or {}
    resolved_year = _select_ess_import_year(year_candidates, fallback=annee)
    if year_check["issues"]:
        for issue in year_check["issues"]:
            print(f"  ⚠ {issue}")
    if resolved_year is None:
        wb.close()
        print("  ✗ Année ESS introuvable : import interrompu.")
        return False
    if annee != resolved_year:
        print(f"  ⚠ Année utilisée pour l'import : {resolved_year}")
        annee = resolved_year

    # ── 1. Enregistrement de la source ──────────────────────────────────────
    if not dry_run and conn:
        removed = purge_existing_ess_import(conn, institution, annee, nom_fichier)
        if removed:
            print(f"  ↺ Ancien import remplacé ({removed} source(s) supprimée(s))")
        source_id = register_source(
            conn,
            type_source      = 'ESS',
            nom_fichier      = nom_fichier,
            chemin_fichier   = chemin_rel,
            institution      = institution,
            annee_donnees    = annee,
            description      = f"ESS OIT/BIT — {institution} {annee}",
            fiabilite        = 'primaire',
            note_methodologique = note_anomalie
        )
    else:
        source_id = 0  # dry_run

    # ── 2. Parse de l'inventaire ─────────────────────────────────────────────
    INVENTAIRE_SHEET = 'Inventaire des régimes'
    if INVENTAIRE_SHEET not in wb.sheetnames:
        print(f"  ✗ Feuille '{INVENTAIRE_SHEET}' absente")
        return False

    ws_inv = wb[INVENTAIRE_SHEET]
    regimes = parse_inventaire(ws_inv, institution, annee, verbose=verbose)
    print(f"  Inventaire : {len(regimes)} régime(s) trouvé(s)")

    regime_map = {}  # num → code (ex: 1 → 'CNSS_R1')
    for r in regimes:
        regime_map[r['regime_num']] = r['regime_code']
        if not dry_run and conn:
            # regimes_historique
            upsert_regime(conn, r['institution'], r['regime_code'], r['annee'], source_id,
                nom_original    = r['nom_original'],
                nom_fr          = r['nom_fr'],
                administrateur  = r['administrateur'],
                type_financement= r['type_financement'],
                caractere       = r['caractere'],
                type_assurance  = r['type_assurance'],
                gestion         = r['gestion'],
                fonctions_oit   = r['fonctions_oit'],
                statut_regime   = 'transitoire' if 'reforme' in (r['nom_fr'] or '').lower()
                                  else 'actif',
            )
            # indicateurs_regime
            upsert_indicateurs(conn, r['institution'], r['regime_code'], r['annee'], source_id,
                cotisants_total          = r['cotisants_total'],
                cotisants_h              = r['cotisants_h'],
                cotisants_f              = r['cotisants_f'],
                beneficiaires_total      = r['beneficiaires_total'],
                beneficiaires_h          = r['beneficiaires_h'],
                beneficiaires_f          = r['beneficiaires_f'],
                recettes_cdf             = r['recettes_cdf'],
                recettes_usd             = r['recettes_usd'],
                depenses_prestations_cdf = r['depenses_prestations_cdf'],
                depenses_admin_cdf       = r['depenses_admin_cdf'],
                unite_monetaire_source   = r['unite_monetaire_source'],
            )

    # ── 3. Parse des feuilles de prestations ────────────────────────────────
    for sheet_name in wb.sheetnames:
        regime_suffix = _resolve_sheet_regime_suffix(sheet_name)
        if not regime_suffix:
            continue

        # Résoudre le code régime pour cette feuille
        regime_code   = f"{institution}_{regime_suffix}"

        ws_p = wb[sheet_name]
        prestations = parse_prestation_sheet(ws_p, institution, regime_code, annee, verbose=verbose)

        if prestations:
            print(f"  [{sheet_name}] → {regime_code} : {len(prestations)} prestation(s)")
        if not dry_run and conn:
            for p in prestations:
                upsert_prestation(conn,
                    p['institution'], p['regime_code'], p['annee'], p['prestation_num'],
                    source_id,
                    nom_original             = p['nom_original'],
                    nom_fr                   = p['nom_fr'],
                    fonction_oit             = p['fonction_oit'],
                    groupe_population        = p['groupe_population'],
                    groupe_age               = p['groupe_age'],
                    zone_geo                 = p['zone_geo'],
                    type_financement         = p['type_financement'],
                    couverture_effective_total = p['couverture_effective_total'],
                    couverture_h             = p['couverture_h'],
                    couverture_f             = p['couverture_f'],
                    beneficiaires_total      = p['beneficiaires_total'],
                    beneficiaires_h          = p['beneficiaires_h'],
                    beneficiaires_f          = p['beneficiaires_f'],
                    type_paiement            = p['type_paiement'],
                    periodicite              = p['periodicite'],
                    montant_unitaire_cdf     = p['montant_unitaire_cdf'],
                    montant_unitaire_usd     = p['montant_unitaire_usd'],
                    critere_eligibilite      = p['critere_eligibilite'],
                    duree_service_requise    = p['duree_service_requise'],
                    age_legal_h              = p['age_legal_h'],
                    age_legal_f              = p['age_legal_f'],
                    condition_complementaire = p['condition_complementaire'],
                    depenses_regime_cdf      = p['depenses_regime_cdf'],
                )

    if not dry_run and conn:
        conn.commit()

    print(f"  ✓ Traitement terminé (source_id={source_id})")
    return True


def process_ess_inbox_file(filepath, conn=None, dry_run=False, verbose=False,
                           institution_filter=None, annee_filter=None):
    """
    Traite un fichier ESS déposé dans le dossier de réception unique.
    Le fichier est d'abord normalisé puis enregistré en base.
    """
    base = os.path.basename(filepath)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=False, data_only=True)
    except Exception as exc:
        print(f"  ✗ Impossible d'ouvrir le classeur : {base} ({exc})")
        return False

    institution = _infer_ess_institution_for_file(filepath, wb)
    institution = _sanitize_institution_code(institution)
    year_check = _resolve_ess_year_consistency(filepath, wb)
    year_candidates = year_check.get("candidates") or {}
    annee = _select_ess_import_year(year_candidates)
    if year_check["issues"]:
        for issue in year_check["issues"]:
            print(f"  ⚠ {issue}")

    if institution_filter and institution and institution.upper() != institution_filter.upper():
        wb.close()
        return None
    if annee_filter and annee and annee != annee_filter:
        wb.close()
        return None

    if not institution or not annee:
        wb.close()
        print(f"  ✗ Métadonnées ESS insuffisantes pour {base} (institution={institution}, année={annee})")
        return False

    destination_dir = _ess_destination_dir(institution)
    if not destination_dir:
        wb.close()
        print(f"  ✗ Impossible de déterminer le dossier destination pour l'institution {institution!r}")
        return False

    destination_name = _ess_destination_name(filepath, institution, annee)
    destination_path = os.path.normpath(os.path.join(destination_dir, destination_name))
    destination_rel = os.path.relpath(destination_path, start=os.path.normpath(os.path.join(_SCRIPT_DIR, '..')))

    wb.close()
    os.makedirs(destination_dir, exist_ok=True)
    if os.path.exists(destination_path) and os.path.abspath(destination_path) != os.path.abspath(filepath):
        os.remove(destination_path)

    ok = process_ess_file(
        filepath,
        institution,
        annee,
        conn=conn,
        dry_run=dry_run,
        verbose=verbose,
        nom_fichier_override=destination_name,
        chemin_rel_override=destination_rel,
    )

    if not ok:
        return False

    if dry_run:
        return True

    if os.path.abspath(filepath) != os.path.abspath(destination_path):
        moved, move_error = _move_file_with_retry(filepath, destination_path)
        if not moved:
            print(f"  ✗ Fichier importé mais non déplacé (verrouillage fichier) : {base}")
            print("    Fermer Excel/aperçu du fichier et laisser OneDrive finir la synchronisation, puis relancer.")
            print(f"    Détail : {move_error}")
            return False
        print(f"  ↳ Déplacé vers {destination_rel}")

    return True


# ---------------------------------------------------------------------------
# Point d'entrée principal
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Extraire les données ESS et les charger dans la base de données.')
    parser.add_argument('--annee',       type=int, help='Traiter seulement cette année')
    parser.add_argument('--institution', type=str, help='Traiter seulement cette institution (CNSS, CNSSAP)')
    parser.add_argument('--nom-fichier', type=str,
                        help='Filtrer sur un nom de fichier source ESS exact (pour suppression)')
    parser.add_argument('--source-id', type=int,
                        help='Identifiant source_id précis (pour suppression)')
    parser.add_argument('--delete', action='store_true',
                        help='Supprimer des imports ESS existants de la base')
    parser.add_argument('--all', action='store_true',
                        help='Autoriser la suppression de toutes les sources ESS (avec --delete)')
    parser.add_argument('--force', action='store_true',
                        help='Confirmer une suppression effective (sans --dry-run)')
    parser.add_argument('--inbox', action='store_true',
                        help='Traiter les fichiers ESS déposés dans 06_sources/_entrants/')
    parser.add_argument('--inbox-dir', type=str, default=ESS_INBOX_DIR,
                        help='Chemin du dossier de réception ESS')
    parser.add_argument('--validate-report', action='store_true',
                        help='Générer un rapport de validation sans écrire en base')
    parser.add_argument('--report-dir', type=str, default=ESS_REPORT_DIR,
                        help='Dossier de sortie des rapports de validation')
    parser.add_argument('--dry-run',     action='store_true',
                        help='Simulation : parse les fichiers sans écrire dans la BDD')
    parser.add_argument('--verbose',     action='store_true',
                        help='Afficher le détail de chaque ligne parsée')
    args = parser.parse_args()

    print("=" * 65)
    print("  Extracteur ESS — Protection sociale RDC")
    print("=" * 65)
    if args.dry_run:
        print("  MODE DRY-RUN : aucune écriture en base")
    if args.delete:
        print("  MODE SUPPRESSION ESS")
    if args.validate_report:
        print("  MODE VALIDATION : rapports sans écriture en base")

    if args.validate_report and args.delete:
        print("✗ La génération de rapport ne peut pas être combinée avec --delete.")
        return

    if args.validate_report:
        resultats = {'ok': 0, 'erreur': 0}
        if args.inbox:
            inbox_files = discover_ess_inbox_files(args.inbox_dir)
            if not inbox_files:
                print(f"Aucun fichier ESS trouvé dans {args.inbox_dir}.")
            for filepath in inbox_files:
                report = validate_ess_workbook(filepath, source_mode="inbox")
                if args.institution and report.get("detected_institution") and report.get("detected_institution").upper() != args.institution.upper():
                    continue
                if args.annee and report.get("detected_year") and report.get("detected_year") != args.annee:
                    continue
                out_path = _write_validation_report(report, args.report_dir)
                print(f"  ↳ Rapport écrit : {out_path}")
                if report.get("verdict") == "blocked":
                    resultats['erreur'] += 1
                else:
                    resultats['ok'] += 1
        else:
            archive_entries, skipped_entries = discover_ess_archive_files(ESS_BASE_DIR)
            if args.annee:
                archive_entries = [entry for entry in archive_entries if entry.get("annee") == args.annee]
                skipped_entries = [entry for entry in skipped_entries if entry.get("annee") == args.annee]
            if args.institution:
                archive_entries = [
                    entry for entry in archive_entries
                    if (entry.get("institution") or "").upper() == args.institution.upper()
                ]
                skipped_entries = [
                    entry for entry in skipped_entries
                    if (entry.get("institution") or "").upper() == args.institution.upper()
                ]
            if skipped_entries:
                print("\n  Fichiers ESS archivés ignorés (doublons ou copies) :")
                for entry in skipped_entries:
                    print(f"    - {entry['rel_path']} — {entry['skip_reason']}")
            if not archive_entries:
                print("Aucun fichier ESS correspondant aux critères dans 06_sources/ESS.")
                return
            for entry in archive_entries:
                report = validate_ess_workbook(
                    entry["filepath"],
                    institution=entry.get("institution"),
                    annee=entry.get("annee"),
                    note_anomalie=entry.get("note_anomalie"),
                    source_mode="archive_scan",
                )
                out_path = _write_validation_report(report, args.report_dir)
                print(f"  ↳ Rapport écrit : {out_path}")
                if report.get("verdict") == "blocked":
                    resultats['erreur'] += 1
                else:
                    resultats['ok'] += 1

        print(f"\n{'='*65}")
        print(f"  Résumé : {resultats['ok']} rapport(s) généré(s), {resultats['erreur']} cas bloquant(s)")
        print()
        return

    # Initialiser / ouvrir la base
    conn = None
    if (not args.dry_run) or args.delete:
        conn = create_or_update_db(verbose=True)

    if args.delete:
        has_filter = any([
            args.source_id is not None,
            bool(args.institution),
            args.annee is not None,
            bool(args.nom_fichier),
        ])
        if not has_filter and not args.all:
            print("✗ Suppression refusée : préciser au moins un filtre (--source-id, --institution, --annee, --nom-fichier) ou --all.")
            if conn:
                conn.close()
            return

        if not args.dry_run and not args.force:
            print("✗ Suppression bloquée : relancer avec --force (ou utiliser --dry-run pour simuler).")
            if conn:
                conn.close()
            return

        matches = find_ess_sources(
            conn,
            institution=args.institution,
            annee=args.annee,
            nom_fichier=args.nom_fichier,
            source_id=args.source_id,
        )
        if not matches:
            print("Aucune source ESS correspondant aux critères.")
            if conn:
                conn.close()
            return

        print(f"\n  Sources ESS ciblées ({len(matches)}) :")
        for row in matches:
            print(f"    - source_id={row['source_id']} | {row['institution']} {row['annee_donnees']} | {row['nom_fichier']}")

        source_ids = [row['source_id'] for row in matches]
        deleted = delete_ess_sources(conn, source_ids, dry_run=args.dry_run)

        action = "seraient supprimées" if args.dry_run else "supprimées"
        print(f"\n  ✓ Sources {action} : {deleted['sources']}")
        print(f"    - regimes_historique : {deleted['regimes']}")
        print(f"    - indicateurs_regime : {deleted['indicateurs']}")
        print(f"    - prestations_historique : {deleted['prestations']}")
        if conn:
            conn.close()
        return

    resultats = {'ok': 0, 'erreur': 0}

    if args.inbox:
        inbox_files = discover_ess_inbox_files(args.inbox_dir)
        if not inbox_files:
            print(f"Aucun fichier ESS trouvé dans {args.inbox_dir}.")
        for filepath in inbox_files:
            ok = process_ess_inbox_file(
                filepath,
                conn=conn,
                dry_run=args.dry_run,
                verbose=args.verbose,
                institution_filter=args.institution,
                annee_filter=args.annee,
            )
            if ok is True:
                resultats['ok'] += 1
            elif ok is False:
                resultats['erreur'] += 1
    else:
        archive_entries, skipped_entries = discover_ess_archive_files(ESS_BASE_DIR)
        if args.annee:
            archive_entries = [entry for entry in archive_entries if entry.get("annee") == args.annee]
            skipped_entries = [entry for entry in skipped_entries if entry.get("annee") == args.annee]
        if args.institution:
            archive_entries = [
                entry for entry in archive_entries
                if (entry.get("institution") or "").upper() == args.institution.upper()
            ]
            skipped_entries = [
                entry for entry in skipped_entries
                if (entry.get("institution") or "").upper() == args.institution.upper()
            ]

        if skipped_entries:
            print("\n  Fichiers ESS archivés ignorés (doublons ou copies) :")
            for entry in skipped_entries:
                print(f"    - {entry['rel_path']} — {entry['skip_reason']}")

        if not archive_entries:
            print("Aucun fichier ESS correspondant aux critères dans 06_sources/ESS.")
            return

        for entry in archive_entries:
            ok = process_ess_file(
                entry["filepath"],
                entry.get("institution"),
                entry.get("annee"),
                note_anomalie=entry.get("note_anomalie"),
                conn=conn,
                dry_run=args.dry_run,
                verbose=args.verbose,
            )
            if ok is True:
                resultats['ok'] += 1
            elif ok is False:
                resultats['erreur'] += 1

    # ── Résumé final ─────────────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print(f"  Résumé : {resultats['ok']} fichier(s) traité(s), {resultats['erreur']} erreur(s)")

    if conn and not args.dry_run:
        # Afficher un aperçu de ce qui a été inséré
        print("\n  Contenu de la base après ingestion :")
        for table in ('sources_ingestion', 'regimes_historique',
                      'indicateurs_regime', 'prestations_historique'):
            n = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            print(f"    {table:35s} : {n:4d} ligne(s)")
        conn.close()

    print()


if __name__ == '__main__':
    main()
