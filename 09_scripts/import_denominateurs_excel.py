"""
import_denominateurs_excel.py
─────────────────────────────
Importe la feuille REF_DB_Denominators du fichier Excel
"Denominateurs 2026.xlsx" dans la table `denominateurs_ref`
de la base SQLite protection_sociale_rdc.db.

Seules les lignes COD (RDC) sont conservées.

Usage :
    py 09_scripts/import_denominateurs_excel.py
"""

import sqlite3
import sys
from pathlib import Path

# ── Chemins ─────────────────────────────────────────────────────────────────

WORKSPACE = Path(__file__).resolve().parent.parent
EXCEL_PATH = WORKSPACE / "06_sources" / "Calculateurs et dénominateurs" / "Denominateurs 2026.xlsx"
DB_PATH    = WORKSPACE / "protection_sociale_rdc.db"
ISO_FILTER = "COD"         # Uniquement la RDC
SHEET_NAME = "REF_DB_Denominators"

# ── DDL ─────────────────────────────────────────────────────────────────────

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS denominateurs_ref (
    id           INTEGER,
    var_code     TEXT    NOT NULL,
    iso3         TEXT    NOT NULL,
    year         INTEGER NOT NULL,
    year_dp      TEXT,
    class_sex    TEXT,
    class_age    TEXT,
    val_n        REAL,
    source       TEXT,
    source_note  TEXT,
    sys_note     TEXT,
    priority     INTEGER DEFAULT 0
);
"""

INDEX_SQL = """
CREATE INDEX IF NOT EXISTS idx_denom_lookup
    ON denominateurs_ref (iso3, var_code, class_sex, class_age, year);
"""

# ── Import ───────────────────────────────────────────────────────────────────

def import_denominateurs():
    if not EXCEL_PATH.exists():
        print(f"[ERREUR] Fichier Excel introuvable : {EXCEL_PATH}")
        sys.exit(1)

    print(f"[INFO] Lecture de : {EXCEL_PATH.name}")

    try:
        import openpyxl
    except ImportError:
        print("[ERREUR] openpyxl non installé. Lancer : pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERREUR] Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    rows_to_insert = []
    total_read = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:          # Fin des données
            break
        total_read += 1
        iso3 = row[2]
        if iso3 != ISO_FILTER:
            continue
        rows_to_insert.append((
            row[0],   # id
            row[1],   # var_code
            row[2],   # iso3
            row[3],   # year
            row[4],   # year_dp
            row[5],   # class_sex
            row[6],   # class_age
            row[7],   # val_n
            row[8],   # source
            row[9],   # source_note
            row[10],  # sys_note
            row[11],  # priority
        ))

    wb.close()
    print(f"[INFO] Lignes lues : {total_read} | Lignes COD retenues : {len(rows_to_insert)}")

    if not rows_to_insert:
        print("[AVERT] Aucune ligne COD trouvée. Import annulé.")
        sys.exit(0)

    # ── SQLite ──────────────────────────────────────────────────────────────
    con = sqlite3.connect(str(DB_PATH))
    cur = con.cursor()

    cur.execute(CREATE_SQL)

    # Supprimer les données COD existantes avant réimport
    cur.execute("DELETE FROM denominateurs_ref WHERE iso3 = ?", (ISO_FILTER,))
    deleted = cur.rowcount
    if deleted:
        print(f"[INFO] {deleted} lignes COD existantes supprimées avant réimport.")

    cur.executemany(
        """INSERT INTO denominateurs_ref
               (id, var_code, iso3, year, year_dp, class_sex, class_age,
                val_n, source, source_note, sys_note, priority)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        rows_to_insert,
    )
    cur.execute(INDEX_SQL)
    con.commit()

    # ── Vérification rapide ──────────────────────────────────────────────────
    cur.execute("SELECT COUNT(*) FROM denominateurs_ref WHERE iso3 = ?", (ISO_FILTER,))
    count = cur.fetchone()[0]
    print(f"[OK] {count} lignes COD enregistrées dans denominateurs_ref.")

    cur.execute("""
        SELECT var_code, class_sex, class_age,
               MIN(year) AS yr_min, MAX(year) AS yr_max, COUNT(*) AS n
        FROM denominateurs_ref
        WHERE iso3 = ?
        GROUP BY var_code, class_sex, class_age
        ORDER BY var_code, class_sex, class_age
    """, (ISO_FILTER,))
    print("\n[RÉSUMÉ DES SÉRIES COD]")
    print(f"  {'var_code':<22} {'sex':<10} {'age':<15} {'années':<15} n")
    print("  " + "-" * 70)
    for r in cur.fetchall():
        vc, sx, ag, y0, y1, n = r
        print(f"  {str(vc):<22} {str(sx):<10} {str(ag):<15} {str(y0)}-{str(y1):<10} {n}")

    con.close()
    print("\n[TERMINÉ]")


if __name__ == "__main__":
    import_denominateurs()
